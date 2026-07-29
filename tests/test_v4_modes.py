"""
test_v4_modes.py - 記述式のみモード機能のテスト

α: 記述採点のリッチ化（DescriptiveReviewGUI, generate_descriptive_only_sheets）
"""

import json
import shutil
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest
import tkinter as tk
import numpy as np

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "main_src"))

from conftest import get_shared_tk_root

from constants import (
    RESULTS_FOLDER,
    RESULTS_DATA_FOLDER,
    BOXED_FOLDER,
    SCORED_FOLDER,
    SESSION_STATE_FILE,
    get_rendering_settings,
)


# ============================================================
# ヘルパー
# ============================================================

def _make_stub_app(img_folder=""):
    """Mark2GUI の軽量スタブ（__init__ をスキップ）"""
    from main_gui import Mark2GUI

    root = get_shared_tk_root()
    app = object.__new__(Mark2GUI)
    app.root = root
    app.image_folder_path = tk.StringVar(root, value=img_folder)
    app.skip_questions = tk.StringVar(root, value="4")
    app.descriptive_enabled = tk.BooleanVar(root, value=True)
    app.rendering_settings = get_rendering_settings()
    app._log_messages = []
    app.log_message = lambda msg: app._log_messages.append(msg)
    app._desc_status_label = tk.Label(root)
    app._desc_status_frame = tk.Frame(root)
    app.name_trim_enabled = tk.BooleanVar(root, value=False)
    app._processing = False
    return app


# ============================================================
# generate_descriptive_only_sheets テスト
# ============================================================

class TestGenerateDescriptiveOnlySheets(unittest.TestCase):
    """記述のみモードの採点済み答案生成"""

    def setUp(self):
        self.tmpdir = Path(tempfile.mkdtemp())
        self.boxed_folder = self.tmpdir / "boxed"
        self.boxed_folder.mkdir()
        self.output_folder = self.tmpdir / "scored"

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _make_test_image(self, name="test001.jpg"):
        """テスト画像（595x842 白画像）を作成"""
        import cv2
        img = np.ones((842, 595, 3), dtype=np.uint8) * 255
        cv2.imwrite(str(self.boxed_folder / name), img)

    def test_empty_folder(self):
        """画像がない場合はカウント0"""
        from descriptive_scorer import generate_descriptive_only_sheets

        config = {"questions": [{"id": "D1", "name": "記述1", "max_score": 5,
                                  "aspect": 1, "region": [10, 10, 200, 100]}]}
        result = generate_descriptive_only_sheets(
            boxed_folder=str(self.boxed_folder),
            config=config,
            descriptive_scores={},
            output_folder=str(self.output_folder),
        )
        self.assertEqual(result["total_count"], 0)

    def test_generates_scored_image(self):
        """画像を入力すると採点済み画像が出力される"""
        from descriptive_scorer import generate_descriptive_only_sheets

        self._make_test_image("student01.jpg")
        config = {
            "questions": [
                {"id": "D1", "name": "記述1", "max_score": 5,
                 "aspect": 1, "region": [50, 50, 300, 200]},
            ],
        }
        scores = {"student01.jpg": {"D1": 3}}

        result = generate_descriptive_only_sheets(
            boxed_folder=str(self.boxed_folder),
            config=config,
            descriptive_scores=scores,
            output_folder=str(self.output_folder),
        )
        self.assertEqual(result["total_count"], 1)
        self.assertEqual(result["success_count"], 1)
        self.assertEqual(result["error_count"], 0)
        self.assertTrue((self.output_folder / "student01.jpg").exists())

    def test_handles_missing_scores(self):
        """スコアがない生徒も処理される"""
        from descriptive_scorer import generate_descriptive_only_sheets

        self._make_test_image("student01.jpg")
        self._make_test_image("student02.jpg")
        config = {
            "questions": [
                {"id": "D1", "name": "記述1", "max_score": 5,
                 "aspect": 1, "region": [50, 50, 300, 200]},
            ],
        }
        scores = {"student01.jpg": {"D1": 3}}  # student02 has no scores

        result = generate_descriptive_only_sheets(
            boxed_folder=str(self.boxed_folder),
            config=config,
            descriptive_scores=scores,
            output_folder=str(self.output_folder),
        )
        self.assertEqual(result["total_count"], 2)
        self.assertEqual(result["success_count"], 2)

    def test_log_callback(self):
        """log_callback が呼ばれる"""
        from descriptive_scorer import generate_descriptive_only_sheets

        self._make_test_image("test.jpg")
        logs = []
        config = {"questions": [{"id": "D1", "name": "Q1", "max_score": 3,
                                  "aspect": 1, "region": [10, 10, 100, 50]}]}

        generate_descriptive_only_sheets(
            boxed_folder=str(self.boxed_folder),
            config=config,
            descriptive_scores={"test.jpg": {"D1": 2}},
            output_folder=str(self.output_folder),
            log_callback=lambda msg: logs.append(msg),
        )
        self.assertTrue(any("記述のみ" in m for m in logs))


# ============================================================
# process_descriptive_only_summary テスト
# ============================================================

class TestProcessDescriptiveOnlySummary(unittest.TestCase):
    """記述のみモードのサマリー生成"""

    def setUp(self):
        self.tmpdir = Path(tempfile.mkdtemp())
        self.img_folder = self.tmpdir / "images"
        self.img_folder.mkdir()
        (self.img_folder / RESULTS_FOLDER).mkdir()

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_basic_summary(self):
        """基本的なサマリー生成"""
        from summary_generator import process_descriptive_only_summary

        config = {
            "questions": [
                {"id": "D1", "name": "記述1", "max_score": 5, "aspect": 1,
                 "region": [10, 10, 200, 100]},
                {"id": "D2", "name": "記述2", "max_score": 3, "aspect": 2,
                 "region": [10, 110, 200, 200]},
            ]
        }
        scores = {
            "student01.jpg": {"D1": 5, "D2": 3},
            "student02.jpg": {"D1": 3, "D2": 1},
            "student03.jpg": {"D1": 2, "D2": 2},
        }

        result = process_descriptive_only_summary(
            image_folder=str(self.img_folder),
            descriptive_config=config,
            descriptive_scores=scores,
        )
        self.assertTrue(result["success"])
        stats = result["stats"]
        self.assertEqual(stats["受験者数"], 3)
        self.assertEqual(stats["満点"], 8)
        self.assertAlmostEqual(stats["平均点"], (8 + 4 + 4) / 3, places=1)
        self.assertEqual(stats["最高点"], 8)
        self.assertEqual(stats["最低点"], 4)

    def test_empty_questions(self):
        """問題が空の場合はエラー"""
        from summary_generator import process_descriptive_only_summary

        result = process_descriptive_only_summary(
            image_folder=str(self.img_folder),
            descriptive_config={"questions": []},
            descriptive_scores={},
        )
        self.assertFalse(result["success"])
        self.assertIn("設定されていません", result["error"])

    def test_excel_files_created(self):
        """Excelファイルが生成される"""
        from summary_generator import process_descriptive_only_summary
        from constants import STUDENT_SUMMARY_FILE, EXAM_SUMMARY_FILE, FINAL_REPORT_FOLDER

        config = {
            "questions": [
                {"id": "D1", "name": "記述1", "max_score": 5, "aspect": 1,
                 "region": [10, 10, 200, 100]},
            ]
        }
        scores = {"s1.jpg": {"D1": 3}}

        result = process_descriptive_only_summary(
            image_folder=str(self.img_folder),
            descriptive_config=config,
            descriptive_scores=scores,
        )
        self.assertTrue(result["success"])

        student_path = Path(result["student_summary_path"])
        exam_path = Path(result["exam_summary_path"])
        self.assertTrue(student_path.exists())
        self.assertTrue(exam_path.exists())

    def test_roster_order_controls_row_order(self):
        """名簿(roster)を渡すと、出力行が名簿の並び順になること"""
        from summary_generator import process_descriptive_only_summary
        from openpyxl import load_workbook

        config = {
            "questions": [
                {"id": "D1", "name": "記述1", "max_score": 5, "aspect": 1,
                 "region": [10, 10, 200, 100]},
            ]
        }
        # ファイル名の自然順とは異なる順で名簿を用意する
        scores = {
            "c.jpg": {"D1": 1},
            "a.jpg": {"D1": 2},
            "b.jpg": {"D1": 3},
        }
        student_id_result = {
            "c.jpg": {"text": "20230002", "name": "佐藤"},
            "a.jpg": {"text": "20230001", "name": "山田"},
            "b.jpg": {"text": "20230003", "name": "鈴木"},
        }
        roster = {"20230003": "鈴木", "20230001": "山田", "20230002": "佐藤"}

        result = process_descriptive_only_summary(
            image_folder=str(self.img_folder),
            descriptive_config=config,
            descriptive_scores=scores,
            student_id_result=student_id_result,
            roster=roster,
        )
        self.assertTrue(result["success"])

        wb = load_workbook(result["student_summary_path"])
        ws = wb.active
        # B列=ファイル名の並びが、名簿の順(鈴木→山田→佐藤 = b,a,c)になっているはず
        filenames_in_order = [ws.cell(row=r, column=2).value for r in (2, 3, 4)]
        self.assertEqual(filenames_in_order, ["b.jpg", "a.jpg", "c.jpg"])

    def test_no_roster_falls_back_to_filename_order(self):
        """名簿を渡さない場合は従来通りファイル名の自然順になること"""
        from summary_generator import process_descriptive_only_summary
        from openpyxl import load_workbook

        config = {
            "questions": [
                {"id": "D1", "name": "記述1", "max_score": 5, "aspect": 1,
                 "region": [10, 10, 200, 100]},
            ]
        }
        scores = {"c.jpg": {"D1": 1}, "a.jpg": {"D1": 2}, "b.jpg": {"D1": 3}}

        result = process_descriptive_only_summary(
            image_folder=str(self.img_folder),
            descriptive_config=config,
            descriptive_scores=scores,
        )
        self.assertTrue(result["success"])

        wb = load_workbook(result["student_summary_path"])
        ws = wb.active
        filenames_in_order = [ws.cell(row=r, column=2).value for r in (2, 3, 4)]
        self.assertEqual(filenames_in_order, ["a.jpg", "b.jpg", "c.jpg"])

    def test_roster_absent_student_gets_blank_row(self):
        """名簿にいるが未提出の学生も、名簿の行数どおり空欄行として出力されること"""
        from summary_generator import process_descriptive_only_summary
        from openpyxl import load_workbook

        config = {
            "questions": [
                {"id": "D1", "name": "記述1", "max_score": 5, "aspect": 1,
                 "region": [10, 10, 200, 100]},
            ]
        }
        # 名簿には3名いるが、20230002(佐藤)は答案が見つからない(未提出)
        scores = {
            "a.jpg": {"D1": 4},
            "c.jpg": {"D1": 2},
        }
        student_id_result = {
            "a.jpg": {"text": "20230001", "name": "山田"},
            "c.jpg": {"text": "20230003", "name": "鈴木"},
        }
        roster = {"20230001": "山田", "20230002": "佐藤", "20230003": "鈴木"}

        result = process_descriptive_only_summary(
            image_folder=str(self.img_folder),
            descriptive_config=config,
            descriptive_scores=scores,
            student_id_result=student_id_result,
            roster=roster,
        )
        self.assertTrue(result["success"])

        wb = load_workbook(result["student_summary_path"])
        ws = wb.active
        headers = [c.value for c in ws[1]]
        sid_col = headers.index("学籍番号(確認済み)") + 1
        name_col = headers.index("氏名候補(名簿照合)") + 1
        total_col = headers.index("合計") + 1
        score_col = headers.index("記述1 (5)") + 1

        # 3行(名簿の人数分)出力され、2行目が未提出の空欄行になっているはず
        self.assertEqual(ws.cell(row=2, column=2).value, "a.jpg")
        self.assertEqual(ws.cell(row=3, column=2).value, "(未提出)")
        self.assertEqual(ws.cell(row=4, column=2).value, "c.jpg")

        # 未提出行でも学籍番号・氏名は名簿から埋まる
        self.assertEqual(ws.cell(row=3, column=sid_col).value, "20230002")
        self.assertEqual(ws.cell(row=3, column=name_col).value, "佐藤")
        # 得点・合計は空欄(None)のまま
        self.assertIsNone(ws.cell(row=3, column=score_col).value)
        self.assertIsNone(ws.cell(row=3, column=total_col).value)

        # 統計は実際に提出された2名分のみで計算される(未提出者を0点として混入させない)
        stats = result["stats"]
        self.assertEqual(stats["受験者数"], 2)
        self.assertAlmostEqual(stats["平均点"], 3.0, places=1)


# ============================================================
# DescriptiveReviewGUI 構造テスト
# ============================================================

class TestDescriptiveReviewGUIStructure(unittest.TestCase):
    """DescriptiveReviewGUI クラスの存在確認"""

    def test_class_exists(self):
        from descriptive_scorer import DescriptiveReviewGUI
        self.assertIsNotNone(DescriptiveReviewGUI)

    def test_class_attributes(self):
        from descriptive_scorer import DescriptiveReviewGUI
        self.assertTrue(hasattr(DescriptiveReviewGUI, 'THUMB_SIZE_DEFAULT'))
        self.assertTrue(hasattr(DescriptiveReviewGUI, 'GRID_COLS'))
        self.assertEqual(DescriptiveReviewGUI.GRID_COLS, 4)


# ============================================================
# 記述のみモード run_scoring 分岐テスト
# ============================================================

class TestDescriptiveOnlyRunScoring(unittest.TestCase):
    """記述のみモードでの run_scoring 分岐"""

    def setUp(self):
        self.tmpdir = Path(tempfile.mkdtemp())
        self.img_folder = self.tmpdir / "images"
        self.img_folder.mkdir()
        results_data = self.img_folder / RESULTS_FOLDER / RESULTS_DATA_FOLDER
        results_data.mkdir(parents=True)

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    @patch("tkinter.messagebox.showerror")
    def test_no_config_shows_error(self, mock_error):
        """記述設定がない場合はエラー"""
        app = _make_stub_app(str(self.img_folder))
        app._check_descriptive_completeness = lambda: (True, 0, 0, [])
        app._set_processing_state = lambda s: None
        app.log_text = MagicMock()

        app._run_scoring_descriptive_only()
        mock_error.assert_called_once()

    @patch("tkinter.messagebox.showerror")
    def test_no_scores_shows_error(self, mock_error):
        """採点データがない場合はエラー"""
        results_data = self.img_folder / RESULTS_FOLDER / RESULTS_DATA_FOLDER
        config = {"version": 1, "questions": [{"id": "D1", "name": "Q1",
                   "max_score": 5, "aspect": 1, "region": [0, 0, 100, 100]}]}
        (results_data / "descriptive_config.json").write_text(
            json.dumps(config), encoding='utf-8'
        )

        app = _make_stub_app(str(self.img_folder))
        app._check_descriptive_completeness = lambda: (True, 0, 0, [])
        app._set_processing_state = lambda s: None
        app.log_text = MagicMock()

        app._run_scoring_descriptive_only()
        mock_error.assert_called_once()


# ============================================================
# 記述のみモード run_summary_generation 分岐テスト
# ============================================================

class TestDescriptiveOnlySummaryGeneration(unittest.TestCase):
    """記述のみモードでの summary generation 分岐"""

    @patch("tkinter.messagebox.showerror")
    def test_no_data_shows_error(self, mock_error):
        """データがない場合はエラー"""
        tmpdir = Path(tempfile.mkdtemp())
        img_folder = tmpdir / "images"
        img_folder.mkdir()
        results_data = img_folder / RESULTS_FOLDER / RESULTS_DATA_FOLDER
        results_data.mkdir(parents=True)

        try:
            app = _make_stub_app(str(img_folder))
            app._check_descriptive_completeness = lambda: (True, 0, 0, [])

            app._run_summary_generation_descriptive_only()
            mock_error.assert_called_once()
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)


# ============================================================
# generate_descriptive_only_sheets output_scale テスト
# ============================================================

class TestDescriptiveOnlySheetsScale(unittest.TestCase):
    """generate_descriptive_only_sheets の output_scale=1.0 検証"""

    def test_no_double_scaling(self):
        """記述のみモードで座標が二重スケーリングされないことを確認"""
        import cv2
        from descriptive_scorer import generate_descriptive_only_sheets

        tmpdir = Path(tempfile.mkdtemp())
        try:
            boxed = tmpdir / "boxed"
            boxed.mkdir()
            output = tmpdir / "output"

            # 2400x3400 の大きな画像を作成（記述のみモードの典型サイズ）
            img = np.zeros((3400, 2400, 3), dtype=np.uint8)
            img[:] = (255, 255, 255)  # 白背景
            cv2.imwrite(str(boxed / "test.jpg"), img)

            config = {
                "questions": [
                    {"id": "D1", "name": "記述1", "max_score": 5, "aspect": 1,
                     "region": [100, 100, 800, 500]},
                ]
            }
            scores = {"test.jpg": {"D1": 5}}

            result = generate_descriptive_only_sheets(
                boxed_folder=str(boxed),
                config=config,
                descriptive_scores=scores,
                output_folder=str(output),
            )
            self.assertEqual(result["success_count"], 1)
            # 出力画像が存在
            self.assertTrue((output / "test.jpg").exists())
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
