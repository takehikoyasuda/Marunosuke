#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
test_student_id_ocr.py — 学籍番号OCR関連モジュールのテスト

テスト内容:
    Part 1: digit_ocr_preprocessing / digit_ocr_recognizer 単体テスト
    Part 2: roster_loader 単体テスト
    Part 3: student_id_ocr（桁分割・StudentIdOcrTrimmer）単体テスト
    Part 4: generate_student_summary の student_id_result 拡張テスト（回帰含む）
"""

import os
import shutil
import sys
import tempfile
import unittest
from pathlib import Path

import numpy as np
import pandas as pd

sys.path.insert(0, str(Path(__file__).parent.parent / "main_src"))

from PIL import Image

from test_name_trimmer import _make_dummy_mark2_files


# ============================================================
# Part 1: digit_ocr_preprocessing / digit_ocr_recognizer
# ============================================================

class TestDigitOcrPreprocessing(unittest.TestCase):
    """preprocess_digit_image() のテスト"""

    def test_01_blank_image_returns_none(self):
        """空欄(白一色)の画像はNoneを返すこと"""
        from digit_ocr_preprocessing import preprocess_digit_image
        blank = np.full((40, 40, 3), 255, dtype=np.uint8)
        self.assertIsNone(preprocess_digit_image(blank))

    def test_02_inked_image_returns_normalized_array(self):
        """インクのある画像はshape(784,)・dtype float32・値域[0,1]で返ること"""
        from digit_ocr_preprocessing import preprocess_digit_image
        img = np.full((40, 40, 3), 255, dtype=np.uint8)
        img[10:30, 10:30] = 0  # 中央に黒い塊 = インク
        result = preprocess_digit_image(img)
        self.assertIsNotNone(result)
        self.assertEqual(result.shape, (784,))
        self.assertEqual(result.dtype, np.float32)
        self.assertGreaterEqual(result.min(), 0.0)
        self.assertLessEqual(result.max(), 1.0)

    def test_03_grayscale_input_accepted(self):
        """グレースケール(2次元)画像でも動作すること"""
        from digit_ocr_preprocessing import preprocess_digit_image
        img = np.full((40, 40), 255, dtype=np.uint8)
        img[10:30, 10:30] = 0
        result = preprocess_digit_image(img)
        self.assertIsNotNone(result)
        self.assertEqual(result.shape, (784,))


class TestLocalDigitOcrRecognizer(unittest.TestCase):
    """LocalDigitOcrRecognizer のテスト"""

    def test_01_import_and_instantiate(self):
        from digit_ocr_recognizer import LocalDigitOcrRecognizer, DigitOcrCandidate
        recognizer = LocalDigitOcrRecognizer()
        self.assertIsNotNone(recognizer)

    def test_02_empty_list_returns_safe_fallback(self):
        """空リストを渡しても例外にならず、value=Noneで返ること"""
        from digit_ocr_recognizer import LocalDigitOcrRecognizer
        recognizer = LocalDigitOcrRecognizer()
        candidate = recognizer.recognize([])
        self.assertIsNone(candidate.value)
        self.assertEqual(candidate.confidence, 0.0)
        self.assertEqual(candidate.per_digit, [])

    def test_03_model_file_exists(self):
        """モデルファイル(resources/digit_classifier.joblib)が配置されていること"""
        from constants import resource_path
        model_path = resource_path("resources/digit_classifier.joblib")
        self.assertTrue(Path(model_path).exists(), f"モデルファイルが見つかりません: {model_path}")

    def test_04_recognize_returns_per_digit_matching_input_length(self):
        """per_digit の長さが入力画像数と一致すること（精度自体はアサートしない）"""
        from digit_ocr_recognizer import LocalDigitOcrRecognizer
        recognizer = LocalDigitOcrRecognizer()
        digit_images = []
        for _ in range(4):
            img = np.full((40, 40, 3), 255, dtype=np.uint8)
            img[10:30, 10:30] = 0
            digit_images.append(img)
        candidate = recognizer.recognize(digit_images)
        self.assertEqual(len(candidate.per_digit), 4)
        for digit_str, conf in candidate.per_digit:
            self.assertIsInstance(digit_str, str)
            self.assertGreaterEqual(conf, 0.0)
            self.assertLessEqual(conf, 1.0)


# ============================================================
# Part 2: roster_loader
# ============================================================

class TestRosterLoader(unittest.TestCase):
    """load_roster() のテスト"""

    def setUp(self):
        self.test_dir = tempfile.mkdtemp(prefix="test_roster_")

    def tearDown(self):
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def test_01_valid_roster(self):
        from roster_loader import load_roster
        df = pd.DataFrame({
            '学籍番号': ['20230001', '20230002'],
            '氏名': ['山田太郎', '佐藤花子'],
        })
        path = os.path.join(self.test_dir, "roster.xlsx")
        df.to_excel(path, index=False)

        roster = load_roster(path)
        self.assertEqual(roster['20230001'], '山田太郎')
        self.assertEqual(roster['20230002'], '佐藤花子')

    def test_02_missing_column_raises(self):
        from roster_loader import load_roster
        df = pd.DataFrame({'学籍番号': ['20230001'], '名前': ['山田太郎']})
        path = os.path.join(self.test_dir, "roster_bad.xlsx")
        df.to_excel(path, index=False)

        with self.assertRaises(ValueError):
            load_roster(path)


# ============================================================
# Part 3: student_id_ocr
# ============================================================

class TestComputeDigitBoxRects(unittest.TestCase):
    """compute_digit_box_rects() のテスト（1つのrect→マス矩形の幾何計算）"""

    def test_01_even_division_count(self):
        from student_id_ocr import compute_digit_box_rects
        rect = (0, 0, 400, 100)
        rects = compute_digit_box_rects(rect, 8, h_margin_frac=0.0, v_margin_frac=0.0)
        self.assertEqual(len(rects), 8)
        for i, (x0, y0, x1, y1) in enumerate(rects):
            self.assertAlmostEqual(x0, i * 50, delta=1)
            self.assertAlmostEqual(x1, (i + 1) * 50, delta=1)
            self.assertEqual(y0, 0)
            self.assertEqual(y1, 100)

    def test_02_margin_shrinks_each_box(self):
        from student_id_ocr import compute_digit_box_rects
        rect = (0, 0, 400, 100)
        rects = compute_digit_box_rects(rect, 4, h_margin_frac=0.1, v_margin_frac=0.2)
        # 1マス幅100 -> 左右に10ずつマージン -> 幅80
        for (x0, y0, x1, y1) in rects:
            self.assertAlmostEqual(x1 - x0, 80, delta=1)
            # 高さ100 -> 上下に20ずつマージン -> 高さ60
            self.assertAlmostEqual(y1 - y0, 60, delta=1)

    def test_03_rect_used_as_is(self):
        """rectはそのまま(left, top, right, bottom)として扱われること"""
        from student_id_ocr import compute_digit_box_rects
        rect = (0, 0, 400, 100)
        rects = compute_digit_box_rects(rect, 2, h_margin_frac=0.0, v_margin_frac=0.0)
        self.assertEqual(rects[0], (0, 0, 200, 100))
        self.assertEqual(rects[1], (200, 0, 400, 100))


class TestRecognizeStudentIdsIntegration(unittest.TestCase):
    """recognize_student_ids() の統合テスト（id_area_config による位置計算経由）"""

    SAMPLE_IMAGE = Path(__file__).parent.parent / "sample_basefile" / "sample_marksheet.jpg"

    def setUp(self):
        self.test_dir = tempfile.mkdtemp(prefix="test_recognize_sid_")
        self.image_folder = os.path.join(self.test_dir, "boxed")
        os.makedirs(self.image_folder, exist_ok=True)
        shutil.copy(str(self.SAMPLE_IMAGE), os.path.join(self.image_folder, "page_001.jpg"))
        self.output_folder = os.path.join(self.test_dir, "out")

    def tearDown(self):
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def test_01_returns_result_per_image_with_thumbnail(self):
        """設定値(割合)だけからマスを切り出し、画像ごとの候補が返ること"""
        from student_id_ocr import recognize_student_ids
        config = {
            "left_frac": 0.10, "top_frac": 0.10,
            "width_frac": 0.30, "height_frac": 0.05,
            "digit_count": 6,
        }
        results = recognize_student_ids(self.image_folder, config, self.output_folder)
        self.assertEqual(len(results), 1)
        entry = results["page_001.jpg"]
        self.assertIsNotNone(entry["thumbnail_path"])
        self.assertTrue(Path(entry["thumbnail_path"]).exists())
        self.assertEqual(len(entry["per_digit"]), 6)
        self.assertGreaterEqual(entry["confidence"], 0.0)

    def test_02_invalid_image_yields_safe_fallback_entry(self):
        """壊れた画像でも例外を投げず、安全なフォールバック値を返すこと"""
        from student_id_ocr import recognize_student_ids
        bad_folder = os.path.join(self.test_dir, "bad")
        os.makedirs(bad_folder, exist_ok=True)
        with open(os.path.join(bad_folder, "broken.jpg"), "wb") as f:
            f.write(b"not an image")
        config = {
            "left_frac": 0.10, "top_frac": 0.10,
            "width_frac": 0.30, "height_frac": 0.05,
            "digit_count": 6,
        }
        results = recognize_student_ids(bad_folder, config, self.output_folder)
        entry = results["broken.jpg"]
        self.assertIsNone(entry["thumbnail_path"])
        self.assertIsNone(entry["text"])
        self.assertEqual(entry["confidence"], 0.0)
        self.assertEqual(entry["per_digit"], [])


class TestStudentIdOcrTrimmerCleanup(unittest.TestCase):
    """StudentIdOcrTrimmer.cleanup() のテスト"""

    def test_01_cleanup_removes_temp_dir(self):
        from student_id_ocr import StudentIdOcrTrimmer
        trimmer = StudentIdOcrTrimmer()
        temp_dir = tempfile.mkdtemp(prefix="test_id_cleanup_")
        trimmer._temp_dir = temp_dir
        self.assertTrue(Path(temp_dir).exists())

        trimmer.cleanup()
        self.assertFalse(Path(temp_dir).exists())
        self.assertIsNone(trimmer._temp_dir)

    def test_02_cleanup_when_no_temp(self):
        from student_id_ocr import StudentIdOcrTrimmer
        trimmer = StudentIdOcrTrimmer()
        trimmer.cleanup()  # 例外が発生しないこと
        self.assertIsNone(trimmer._temp_dir)


# ============================================================
# Part 4: generate_student_summary の student_id_result 拡張テスト
# ============================================================

def _make_dummy_student_id_result(tmpdir, filenames, with_names=False):
    """テスト用のダミー学籍番号OCR結果(確認済み)を生成する。"""
    result = {}
    img_dir = os.path.join(tmpdir, "student_id_ocr_tmp")
    os.makedirs(img_dir, exist_ok=True)
    for i, filename in enumerate(filenames):
        thumb_path = os.path.join(img_dir, filename)
        Image.new('RGB', (120, 30), color=(220, 220, 200)).save(thumb_path)
        entry = {
            'thumbnail_path': thumb_path,
            'text': f"{20230000 + i}",
            'confidence': 0.9,
        }
        if with_names:
            entry['name'] = f"テスト太郎{i}"
        result[filename] = entry
    return result


class TestGenerateStudentSummaryWithStudentIdResult(unittest.TestCase):
    """generate_student_summary の student_id_result 拡張テスト"""

    @classmethod
    def setUpClass(cls):
        cls.test_dir = tempfile.mkdtemp(prefix="test_summary_sid_")
        cls.template_path, cls.result_path = _make_dummy_mark2_files(
            cls.test_dir, n_students=5, n_questions=3, skip_questions=2
        )
        cls.filenames = [f"page_{i+1:03d}.jpg" for i in range(5)]
        cls.student_id_result = _make_dummy_student_id_result(cls.test_dir, cls.filenames)
        cls.student_id_result_with_names = _make_dummy_student_id_result(
            cls.test_dir, cls.filenames, with_names=True
        )

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.test_dir, ignore_errors=True)

    def test_01_without_student_id_result_regression(self):
        """student_id_result=None で従来通りの列構成になること（回帰テスト）"""
        from saitensamurai import generate_student_summary
        import openpyxl

        output_path = os.path.join(self.test_dir, "summary_no_sid.xlsx")
        df = generate_student_summary(
            self.template_path, self.result_path, output_path,
            skip_questions=2, name_images=None, student_id_result=None,
        )
        self.assertIsNotNone(df)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        headers = [cell.value for cell in ws[2]]
        self.assertEqual(headers[2], '学籍番号1')
        self.assertEqual(ws.freeze_panes, 'C3')
        wb.close()

    def test_02_with_student_id_result_columns_added(self):
        """student_id_result 付きで2列追加されること（画像列＋確認済み列）"""
        from saitensamurai import generate_student_summary
        import openpyxl

        output_path = os.path.join(self.test_dir, "summary_sid.xlsx")
        generate_student_summary(
            self.template_path, self.result_path, output_path,
            skip_questions=2, name_images=None, student_id_result=self.student_id_result,
        )

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        headers = [cell.value for cell in ws[2]]
        self.assertEqual(headers[2], '学籍番号欄')
        self.assertEqual(headers[3], '学籍番号(確認済み)')
        self.assertEqual(headers[4], '学籍番号1')  # OMR方式の列は2つ後ろにずれる
        self.assertEqual(ws.freeze_panes, 'E3')
        wb.close()

    def test_03_with_roster_names_extra_column(self):
        """名簿照合の氏名がある場合、氏名候補列も追加されること"""
        from saitensamurai import generate_student_summary
        import openpyxl

        output_path = os.path.join(self.test_dir, "summary_sid_roster.xlsx")
        generate_student_summary(
            self.template_path, self.result_path, output_path,
            skip_questions=2, name_images=None,
            student_id_result=self.student_id_result_with_names,
        )

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        headers = [cell.value for cell in ws[2]]
        self.assertEqual(headers[2], '学籍番号欄')
        self.assertEqual(headers[3], '学籍番号(確認済み)')
        self.assertEqual(headers[4], '氏名候補(名簿照合)')
        self.assertEqual(headers[5], '学籍番号1')
        self.assertEqual(ws.freeze_panes, 'F3')
        wb.close()

    def test_04_combined_with_name_images(self):
        """氏名欄画像＋学籍番号OCRを併用した場合の列順序が正しいこと"""
        from saitensamurai import generate_student_summary
        import openpyxl

        name_images = {}
        img_dir = os.path.join(self.test_dir, "name_tmp2")
        os.makedirs(img_dir, exist_ok=True)
        for f in self.filenames:
            p = os.path.join(img_dir, f)
            Image.new('RGB', (120, 30), color=(200, 220, 255)).save(p)
            name_images[f] = p

        output_path = os.path.join(self.test_dir, "summary_combined.xlsx")
        generate_student_summary(
            self.template_path, self.result_path, output_path,
            skip_questions=2, name_images=name_images,
            student_id_result=self.student_id_result,
        )

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        headers = [cell.value for cell in ws[2]]
        self.assertEqual(headers[2], '氏名欄')
        self.assertEqual(headers[3], '学籍番号欄')
        self.assertEqual(headers[4], '学籍番号(確認済み)')
        self.assertEqual(headers[5], '学籍番号1')
        self.assertEqual(ws.freeze_panes, 'F3')
        wb.close()

    def test_05_images_embedded(self):
        """学籍番号欄列に画像が埋め込まれていること"""
        from saitensamurai import generate_student_summary
        import openpyxl

        output_path = os.path.join(self.test_dir, "summary_sid_images.xlsx")
        generate_student_summary(
            self.template_path, self.result_path, output_path,
            skip_questions=2, name_images=None, student_id_result=self.student_id_result,
        )

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        self.assertEqual(len(ws._images), 5)
        wb.close()

    def test_06_empty_dict_regression(self):
        """空のstudent_id_result辞書の場合、従来通りの動作になること"""
        from saitensamurai import generate_student_summary
        import openpyxl

        output_path = os.path.join(self.test_dir, "summary_sid_empty.xlsx")
        generate_student_summary(
            self.template_path, self.result_path, output_path,
            skip_questions=2, name_images=None, student_id_result={},
        )

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        headers = [cell.value for cell in ws[2]]
        self.assertEqual(headers[2], '学籍番号1')
        self.assertEqual(ws.freeze_panes, 'C3')
        wb.close()


if __name__ == '__main__':
    unittest.main()
