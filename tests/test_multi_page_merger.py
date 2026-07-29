#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
test_multi_page_merger.py — multi_page_merger.py（複数ページ答案の統合ツール）のテスト。

process_descriptive_only_summary が出力する「記述のみモード」の集計Excelと
同じ列構成を openpyxl で組み立て、read_page_summary / merge_page_summaries /
write_merged_excel の各ロジック（GUI非依存）を検証する。
"""

import shutil
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent / "main_src"))


def _build_summary_excel(path, questions, students, include_sid_column=True):
    """記述のみモードの集計Excelと同じ列構成のダミーファイルを作る。

    Args:
        questions: [(name, max_score), ...]
        students: [{'file': str, 'sid': str|None, 'scores': {name: score}}]
        include_sid_column: '学籍番号(確認済み)' 列を含めるか
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "学生別サマリー"

    headers = ["No.", "ファイル名"]
    if include_sid_column:
        headers += ["学籍番号欄", "学籍番号(確認済み)"]
    for name, max_score in questions:
        headers.append(f"{name} ({max_score})")
    headers.append("合計")
    full_score = sum(m for _, m in questions)
    headers.append(f"配点計 ({full_score})")
    ws.append(headers)

    for i, student in enumerate(students, 1):
        row = [i, student['file']]
        if include_sid_column:
            row += ['', student.get('sid')]
        total = 0
        for name, _ in questions:
            score = student['scores'].get(name, 0)
            row.append(score)
            total += score
        row.append(total)
        ws.append(row)

    wb.save(path)


class TestReadPageSummary(unittest.TestCase):
    def setUp(self):
        self.test_dir = tempfile.mkdtemp(prefix="test_merger_")

    def tearDown(self):
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def test_01_extracts_question_columns_and_rows(self):
        from multi_page_merger import read_page_summary
        path = Path(self.test_dir) / "page1.xlsx"
        _build_summary_excel(
            path,
            questions=[("問1", 10), ("問2", 15)],
            students=[
                {'file': 'a.jpg', 'sid': '1001', 'scores': {'問1': 8, '問2': 12}},
                {'file': 'b.jpg', 'sid': '1002', 'scores': {'問1': 10, '問2': 15}},
            ],
        )
        page = read_page_summary(str(path), "ページ1")
        self.assertEqual(page.question_columns, ["問1 (10)", "問2 (15)"])
        self.assertEqual(page.unmatched_count, 0)
        self.assertEqual(page.max_score, 25.0)
        self.assertEqual(page.rows_by_student_id["1001"]["問1 (10)"], 8)
        self.assertEqual(page.rows_by_student_id["1001"]["合計"], 20)

    def test_02_counts_unmatched_blank_student_id(self):
        from multi_page_merger import read_page_summary
        path = Path(self.test_dir) / "page1.xlsx"
        _build_summary_excel(
            path,
            questions=[("問1", 10)],
            students=[
                {'file': 'a.jpg', 'sid': '1001', 'scores': {'問1': 8}},
                {'file': 'b.jpg', 'sid': None, 'scores': {'問1': 5}},
                {'file': 'c.jpg', 'sid': '', 'scores': {'問1': 3}},
            ],
        )
        page = read_page_summary(str(path), "ページ1")
        self.assertEqual(page.unmatched_count, 2)
        self.assertEqual(len(page.rows_by_student_id), 1)

    def test_03_raises_when_no_student_id_column(self):
        from multi_page_merger import read_page_summary
        path = Path(self.test_dir) / "page1.xlsx"
        _build_summary_excel(
            path, questions=[("問1", 10)],
            students=[{'file': 'a.jpg', 'sid': '1001', 'scores': {'問1': 8}}],
            include_sid_column=False,
        )
        with self.assertRaises(ValueError):
            read_page_summary(str(path), "ページ1")


class TestAnswerPageAudit(unittest.TestCase):
    def setUp(self):
        self.test_dir = Path(tempfile.mkdtemp(prefix="test_page_folders_"))

    def tearDown(self):
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def _answer(self, image_id, exam_page, student_id="", batch_id="batch-1"):
        from multi_page_merger import AnswerPage
        return AnswerPage(
            image_id=image_id,
            image_path=str(self.test_dir / f"{image_id}.png"),
            exam_page=exam_page,
            batch_id=batch_id,
            source_path=str(self.test_dir / f"{batch_id}.pdf"),
            source_page=1,
            student_id=student_id,
        )

    def test_complete_pages_are_associated_by_student_id(self):
        from multi_page_merger import audit_answer_pages
        answers = [
            self._answer("a", 1, "1001"),
            self._answer("b", 1, "1002"),
            self._answer("x", 2, "1002"),
            self._answer("y", 2, "1001"),
        ]

        audit = audit_answer_pages(answers)

        self.assertTrue(audit.is_ready)
        self.assertEqual(audit.associations["1001"], {
            "1": "a", "2": "y",
        })

    def test_missing_duplicate_and_unmatched_are_reported(self):
        from multi_page_merger import audit_answer_pages
        answers = [
            self._answer("a", 1, "1001"),
            self._answer("b", 1, "1001", batch_id="batch-2"),
            self._answer("blank", 1),
            self._answer("x", 2, "1002"),
        ]

        audit = audit_answer_pages(
            answers, expected_pages=[1, 2], roster={"1001": "山田", "1002": "佐藤"},
        )

        self.assertFalse(audit.is_ready)
        self.assertEqual(audit.duplicates["1"]["1001"], ["a", "b"])
        self.assertEqual(audit.unmatched_image_ids, ["blank"])
        self.assertEqual(audit.missing_pages["1001"], [1, 2])
        self.assertEqual(audit.missing_pages["1002"], [1])

    def test_roster_detects_absent_and_unknown_students(self):
        from multi_page_merger import audit_answer_pages
        answers = [
            self._answer("a", 1, "1001"),
            self._answer("c", 1, "9999"),
            self._answer("b", 2, "1001"),
        ]

        audit = audit_answer_pages(
            answers, expected_pages=[1, 2], roster={"1001": "山田", "1002": "佐藤"},
        )

        self.assertEqual(audit.missing_pages["1002"], [1, 2])
        self.assertEqual(audit.unexpected_student_ids, ["9999"])

    def test_batch_accepts_multiple_pdfs_for_same_exam_page(self):
        from multi_page_merger import create_import_batch
        batch = create_import_batch(2, {
            "scan-a.pdf": ["a-1.png", "a-2.png"],
            "scan-b.pdf": ["b-1.png"],
        }, batch_id="batch-1")

        self.assertEqual(batch.exam_page, 2)
        self.assertEqual(batch.source_paths, ["scan-a.pdf", "scan-b.pdf"])
        self.assertEqual(len(batch.answer_pages), 3)
        self.assertEqual(
            [(a.source_path, a.source_page) for a in batch.answer_pages],
            [("scan-a.pdf", 1), ("scan-a.pdf", 2), ("scan-b.pdf", 1)],
        )

    def test_image_files_are_copied_into_managed_batch(self):
        from PIL import Image
        from multi_page_merger import import_files_as_batch
        source1 = self.test_dir / "scan-a.png"
        source2 = self.test_dir / "scan-b.jpg"
        Image.new("RGB", (4, 4), "red").save(source1)
        Image.new("RGB", (4, 4), "blue").save(source2)
        managed = self.test_dir / "managed"

        batch = import_files_as_batch(
            2, [str(source1), str(source2)], str(managed), batch_id="batch-1",
        )

        self.assertEqual(len(batch.answer_pages), 2)
        self.assertTrue(all(Path(answer.image_path).is_file() for answer in batch.answer_pages))
        self.assertTrue(source1.is_file())
        self.assertTrue(source2.is_file())
        self.assertTrue(all(answer.exam_page == 2 for answer in batch.answer_pages))

    def test_failed_import_removes_partial_batch_only(self):
        from PIL import Image
        from multi_page_merger import import_files_as_batch
        source = self.test_dir / "scan.png"
        unsupported = self.test_dir / "notes.txt"
        Image.new("RGB", (4, 4)).save(source)
        unsupported.write_text("not an image", encoding="utf-8")
        managed = self.test_dir / "managed"

        with self.assertRaises(ValueError):
            import_files_as_batch(
                1, [str(source), str(unsupported)], str(managed), batch_id="failed",
            )

        self.assertFalse((managed / "failed").exists())
        self.assertTrue(source.exists())

    def test_selected_exam_page_is_prepared_for_existing_scoring_flow(self):
        from PIL import Image
        from multi_page_merger import (
            prepare_exam_page_workspace, resolve_multi_page_project_folder,
        )
        project = self.test_dir / "project"
        project.mkdir()
        source1 = self.test_dir / "p1.png"
        source2 = self.test_dir / "p2.png"
        Image.new("RGB", (4, 4), "red").save(source1)
        Image.new("RGB", (4, 4), "blue").save(source2)
        answers = [
            self._answer("answer-1", 1, "1001"),
            self._answer("answer-2", 2, "1001"),
        ]
        answers[0].image_path = str(source1)
        answers[1].image_path = str(source2)

        workspace = prepare_exam_page_workspace(
            answers, 2, str(project / "_multi_page_pages"), str(project),
        )

        images = list(Path(workspace).glob("*.png"))
        self.assertEqual([path.name for path in images], ["answer-2.png"])
        self.assertEqual(resolve_multi_page_project_folder(workspace), str(project.resolve()))

    def test_shared_layout_is_published_and_applied_to_another_page(self):
        from constants import RESULTS_DATA_FOLDER, RESULTS_FOLDER
        from multi_page_merger import (
            apply_shared_layout_settings, prepare_exam_page_workspace,
            publish_shared_layout_settings, shared_layout_was_applied,
        )
        project = self.test_dir / "project"
        project.mkdir()
        source1 = self.test_dir / "p1.png"
        source2 = self.test_dir / "p2.png"
        source1.write_bytes(b"page1")
        source2.write_bytes(b"page2")
        answers = [self._answer("a", 1), self._answer("b", 2)]
        answers[0].image_path = str(source1)
        answers[1].image_path = str(source2)
        root = project / "_multi_page_pages"

        page1 = Path(prepare_exam_page_workspace(
            answers, 1, str(root), str(project), shared_layout=True,
        ))
        page1_data = page1 / RESULTS_FOLDER / RESULTS_DATA_FOLDER
        page1_data.mkdir(parents=True)
        (page1_data / "descriptive_config.json").write_text(
            '{"questions": []}', encoding="utf-8",
        )
        (page1_data / "name_area_config.json").write_text(
            '{"rect_frac": [0.1, 0.1, 0.2, 0.2]}', encoding="utf-8",
        )
        copied = publish_shared_layout_settings(str(page1))
        self.assertIn("descriptive_config.json", copied)

        page2 = Path(prepare_exam_page_workspace(
            answers, 2, str(root), str(project), shared_layout=True,
        ))
        page2_data = page2 / RESULTS_FOLDER / RESULTS_DATA_FOLDER
        self.assertTrue((page2_data / "descriptive_config.json").exists())
        self.assertTrue((page2_data / "name_area_config.json").exists())
        self.assertTrue(shared_layout_was_applied(str(page2)))
        self.assertEqual(
            apply_shared_layout_settings(str(project), str(page2)),
            ["descriptive_config.json", "name_area_config.json"],
        )

    def test_shared_layout_bootstraps_from_existing_page_workspace(self):
        from constants import RESULTS_DATA_FOLDER, RESULTS_FOLDER
        from multi_page_merger import (
            apply_shared_layout_settings, bootstrap_shared_layout_settings,
        )
        project = self.test_dir / "project"
        page1_data = (
            project / "_multi_page_pages" / "page_001"
            / RESULTS_FOLDER / RESULTS_DATA_FOLDER
        )
        page1_data.mkdir(parents=True)
        (page1_data / "descriptive_config.json").write_text(
            '{"questions": [{"id": "D1"}]}', encoding="utf-8",
        )
        page2 = project / "_multi_page_pages" / "page_002"

        copied = bootstrap_shared_layout_settings(
            str(project), str(project / "_multi_page_pages"),
        )
        applied = apply_shared_layout_settings(str(project), str(page2))

        self.assertEqual(copied, ["descriptive_config.json"])
        self.assertEqual(applied, ["descriptive_config.json"])
        self.assertTrue(
            (page2 / RESULTS_FOLDER / RESULTS_DATA_FOLDER / "descriptive_config.json").exists()
        )

    def test_all_page_summaries_are_combined_automatically(self):
        from constants import FINAL_REPORT_FOLDER, RESULTS_DATA_FOLDER, RESULTS_FOLDER, STUDENT_SUMMARY_FILE
        from multi_page_merger import (
            MultiPageAudit, create_import_batch, generate_combined_multi_page_summary,
            save_multi_page_manifest,
        )
        project = self.test_dir / "project"
        manifest = project / RESULTS_FOLDER / RESULTS_DATA_FOLDER / "multi_page_manifest.json"
        batch1 = create_import_batch(1, {"p1.pdf": ["p1-a.png"]}, batch_id="b1")
        batch2 = create_import_batch(2, {"p2.pdf": ["p2-a.png"]}, batch_id="b2")
        audit = MultiPageAudit(exam_pages=[1, 2], active_exam_page=2)
        save_multi_page_manifest([batch1, batch2], audit, str(manifest))
        for page, question, score in [(1, "問題1", 8), (2, "問題2", 12)]:
            output = (
                project / "_multi_page_pages" / f"page_{page:03d}"
                / RESULTS_FOLDER / FINAL_REPORT_FOLDER / STUDENT_SUMMARY_FILE
            )
            output.parent.mkdir(parents=True)
            _build_summary_excel(
                output, [(question, 10 if page == 1 else 15)],
                [{'file': f'p{page}-a.png', 'sid': '1001', 'scores': {question: score}}],
            )

        result = generate_combined_multi_page_summary(str(project))

        self.assertTrue(result["success"])
        self.assertEqual(result["student_count"], 1)
        workbook = load_workbook(result["output_path"], data_only=True)
        headers = [cell.value for cell in workbook.active[1]]
        self.assertIn("ページ1: 問題1 (10)", headers)
        self.assertIn("ページ2: 問題2 (15)", headers)
        self.assertIn("全ページ試験統計", workbook.sheetnames)
        stats = {
            row[0].value: row[1].value
            for row in workbook["全ページ試験統計"].iter_rows(min_row=2)
        }
        self.assertEqual(stats["満点"], 25)
        self.assertEqual(stats["平均点"], 20)

    def test_combined_summary_reports_all_pages_missing_student_id(self):
        """学籍番号OCR未確認のページが複数ある場合、最初の1件で止めず
        全ページ分をまとめて報告する（missing_student_id_pages）。"""
        from constants import FINAL_REPORT_FOLDER, RESULTS_DATA_FOLDER, RESULTS_FOLDER, STUDENT_SUMMARY_FILE
        from multi_page_merger import (
            MultiPageAudit, create_import_batch, generate_combined_multi_page_summary,
            save_multi_page_manifest,
        )
        project = self.test_dir / "project"
        manifest = project / RESULTS_FOLDER / RESULTS_DATA_FOLDER / "multi_page_manifest.json"
        batch1 = create_import_batch(1, {"p1.pdf": ["p1-a.png"]}, batch_id="b1")
        batch2 = create_import_batch(2, {"p2.pdf": ["p2-a.png"]}, batch_id="b2")
        batch3 = create_import_batch(3, {"p3.pdf": ["p3-a.png"]}, batch_id="b3")
        audit = MultiPageAudit(exam_pages=[1, 2, 3], active_exam_page=3)
        save_multi_page_manifest([batch1, batch2, batch3], audit, str(manifest))
        for page, include_sid in [(1, False), (2, True), (3, False)]:
            output = (
                project / "_multi_page_pages" / f"page_{page:03d}"
                / RESULTS_FOLDER / FINAL_REPORT_FOLDER / STUDENT_SUMMARY_FILE
            )
            output.parent.mkdir(parents=True)
            _build_summary_excel(
                output, [(f"問題{page}", 10)],
                [{'file': f'p{page}-a.png', 'sid': '1001', 'scores': {f"問題{page}": 5}}],
                include_sid_column=include_sid,
            )

        result = generate_combined_multi_page_summary(str(project))

        self.assertFalse(result["success"])
        self.assertEqual(result["missing_student_id_pages"], [1, 3])

    def test_combined_summary_waits_for_unfinished_pages(self):
        from constants import RESULTS_DATA_FOLDER, RESULTS_FOLDER
        from multi_page_merger import (
            MultiPageAudit, create_import_batch, generate_combined_multi_page_summary,
            save_multi_page_manifest,
        )
        project = self.test_dir / "project"
        manifest = project / RESULTS_FOLDER / RESULTS_DATA_FOLDER / "multi_page_manifest.json"
        batches = [
            create_import_batch(1, {"p1.pdf": ["a.png"]}, batch_id="b1"),
            create_import_batch(2, {"p2.pdf": ["b.png"]}, batch_id="b2"),
        ]
        save_multi_page_manifest(
            batches, MultiPageAudit(exam_pages=[1, 2]), str(manifest),
        )

        result = generate_combined_multi_page_summary(str(project))

        self.assertFalse(result["success"])
        self.assertEqual(result["pending_pages"], [1, 2])

    def test_manifest_round_trip_preserves_batches_and_audit(self):
        from multi_page_merger import (
            audit_answer_pages, create_import_batch, load_multi_page_manifest,
            save_multi_page_manifest,
        )
        batch = create_import_batch(
            1, {"scan.pdf": ["a.png"]}, batch_id="batch-1",
        )
        batch.answer_pages[0].student_id = "1001"
        audit = audit_answer_pages(batch.answer_pages, expected_pages=[1])
        path = self.test_dir / "manifest.json"

        save_multi_page_manifest([batch], audit, str(path))
        loaded_batches, loaded_audit = load_multi_page_manifest(str(path))

        self.assertEqual(loaded_batches[0], batch)
        self.assertEqual(loaded_audit, audit)
        self.assertTrue(loaded_audit.is_ready)

    def test_invalid_expected_pages_are_rejected(self):
        from multi_page_merger import audit_answer_pages
        with self.assertRaises(ValueError):
            audit_answer_pages([], expected_pages=[1, 1])

    def test_answer_assigned_to_unexpected_page_is_reported(self):
        from multi_page_merger import audit_answer_pages
        answer = self._answer("a", 3, "1001")
        audit = audit_answer_pages([answer], expected_pages=[1, 2])
        self.assertFalse(audit.is_ready)
        self.assertEqual(audit.unexpected_exam_pages, {"a": 3})


class TestMergePageSummaries(unittest.TestCase):
    def setUp(self):
        self.test_dir = tempfile.mkdtemp(prefix="test_merger_")

    def tearDown(self):
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def _read(self, filename, questions, students):
        path = Path(self.test_dir) / filename
        _build_summary_excel(path, questions, students)
        from multi_page_merger import read_page_summary
        label = filename.replace('.xlsx', '')
        return read_page_summary(str(path), label)

    def test_01_all_pages_complete_no_missing(self):
        from multi_page_merger import merge_page_summaries
        p1 = self._read("p1.xlsx", [("問1", 10)], [
            {'file': 'a.jpg', 'sid': '1001', 'scores': {'問1': 8}},
        ])
        p2 = self._read("p2.xlsx", [("問2", 15)], [
            {'file': 'a.jpg', 'sid': '1001', 'scores': {'問2': 12}},
        ])
        merged, warnings = merge_page_summaries([p1, p2])
        self.assertEqual(warnings, [])
        row = merged["1001"]
        self.assertEqual(row["p1: 問1 (10)"], 8)
        self.assertEqual(row["p2: 問2 (15)"], 12)
        self.assertEqual(row["総合計"], 20)
        self.assertEqual(row["総配点"], 25.0)
        self.assertEqual(row["欠落ページ"], "")

    def test_02_missing_page_flagged_and_excluded_from_total(self):
        from multi_page_merger import merge_page_summaries
        p1 = self._read("p1.xlsx", [("問1", 10)], [
            {'file': 'a.jpg', 'sid': '1001', 'scores': {'問1': 8}},
            {'file': 'b.jpg', 'sid': '1002', 'scores': {'問1': 5}},
        ])
        p2 = self._read("p2.xlsx", [("問2", 15)], [
            {'file': 'a.jpg', 'sid': '1001', 'scores': {'問2': 12}},
            # 1002はページ2を提出していない
        ])
        merged, warnings = merge_page_summaries([p1, p2])
        self.assertEqual(warnings, [])

        complete_row = merged["1001"]
        self.assertEqual(complete_row["欠落ページ"], "")
        self.assertEqual(complete_row["総合計"], 20)

        incomplete_row = merged["1002"]
        self.assertEqual(incomplete_row["欠落ページ"], "p2")
        self.assertIsNone(incomplete_row["p2: 問2 (15)"])
        self.assertEqual(incomplete_row["総合計"], 5)  # p1の得点のみ

    def test_03_inconsistent_max_score_yields_none_total(self):
        from multi_page_merger import PageSummary, merge_page_summaries
        p1 = PageSummary(
            label="p1", source_path="p1.xlsx", question_columns=["問1 (10)"],
            rows_by_student_id={"1001": {"問1 (10)": 8, "合計": 8}},
            max_score=10.0, unmatched_count=0,
        )
        p2 = PageSummary(
            label="p2", source_path="p2.xlsx", question_columns=["問2 (15)"],
            rows_by_student_id={"1001": {"問2 (15)": 12, "合計": 12}},
            max_score=None, unmatched_count=0,  # 配点計の抽出に失敗したケース
        )
        merged, _ = merge_page_summaries([p1, p2])
        self.assertIsNone(merged["1001"]["総配点"])
        self.assertEqual(merged["1001"]["総合計"], 20)

    def test_04_unmatched_rows_reported_as_warning(self):
        from multi_page_merger import merge_page_summaries
        p1 = self._read("p1.xlsx", [("問1", 10)], [
            {'file': 'a.jpg', 'sid': '1001', 'scores': {'問1': 8}},
            {'file': 'b.jpg', 'sid': None, 'scores': {'問1': 5}},
        ])
        p2 = self._read("p2.xlsx", [("問2", 15)], [
            {'file': 'a.jpg', 'sid': '1001', 'scores': {'問2': 12}},
        ])
        _, warnings = merge_page_summaries([p1, p2])
        self.assertEqual(len(warnings), 1)
        self.assertIn("p1", warnings[0])
        self.assertIn("1件", warnings[0])


class TestWriteMergedExcel(unittest.TestCase):
    def setUp(self):
        self.test_dir = tempfile.mkdtemp(prefix="test_merger_")

    def tearDown(self):
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def test_01_output_has_expected_columns_and_highlight(self):
        from multi_page_merger import PageSummary, write_merged_excel

        p1 = PageSummary(
            label="ページ1", source_path="p1.xlsx", question_columns=["問1 (10)"],
            rows_by_student_id={}, max_score=10.0, unmatched_count=0,
        )
        p2 = PageSummary(
            label="ページ2", source_path="p2.xlsx", question_columns=["問2 (15)"],
            rows_by_student_id={}, max_score=15.0, unmatched_count=0,
        )
        merged_rows = {
            "1001": {
                "ページ1: 問1 (10)": 8, "ページ1 合計": 8,
                "ページ2: 問2 (15)": 12, "ページ2 合計": 12,
                "総合計": 20, "総配点": 25.0, "欠落ページ": "",
            },
            "1002": {
                "ページ1: 問1 (10)": 5, "ページ1 合計": 5,
                "ページ2: 問2 (15)": None, "ページ2 合計": None,
                "総合計": 5, "総配点": 25.0, "欠落ページ": "ページ2",
            },
        }
        output_path = str(Path(self.test_dir) / "merged.xlsx")
        write_merged_excel(merged_rows, [p1, p2], output_path)

        wb = load_workbook(output_path)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        self.assertEqual(
            headers,
            ["No", "学籍番号", "ページ1: 問1 (10)", "ページ1 合計",
             "ページ2: 問2 (15)", "ページ2 合計", "総合計", "総配点", "欠落ページ"],
        )

        # 1001(欠落なし)は2行目、1002(欠落あり)は3行目(merged_rowsの挿入順)
        row2 = [c.value for c in ws[2]]
        row3 = [c.value for c in ws[3]]
        self.assertEqual(row2[1], "1001")
        self.assertEqual(row3[1], "1002")

        # 欠落ページがある行は背景色がついていること
        self.assertEqual(ws.cell(row=3, column=2).fill.fgColor.rgb, "00FFCDD2")
        self.assertNotEqual(ws.cell(row=2, column=2).fill.fgColor.rgb, "00FFCDD2")


if __name__ == '__main__':
    unittest.main()
