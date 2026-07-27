"""
OMR認識エンジン (omr_engine.py)

Mark2のOMR (Optical Mark Recognition) 処理を担うモジュール。
座標管理・マーク認識・認識結果のExcel出力など、マーク採点パイプライン全体を提供する。
画像読み込み・コーナーマーカー検出・射影変換補正は image_alignment.py に切り出し済み
(採点モードを問わず共通で使うため)。

主な機能:
- parse_excel_coordinates: 座標定義Excelのパース
- recognize_marks: マーク認識 (Mark2OSSロジック準拠)
- save_recognition_results: 認識結果のExcel出力
- process_box_drawer: フォルダ一括処理 (枠描画 + OMR認識)
"""

import csv
import json
import logging
import os
import time
import sys
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from pathlib import Path
from datetime import datetime

import cv2
import pandas as pd
import numpy as np
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from constants import (
    RESULTS_FOLDER,
    BOXED_FOLDER,
    CLEAN_FOLDER,
    RESULTS_DATA_FOLDER,
    ANSWER_KEY_FILE,
    READING_RESULTS_FOLDER_NAME,
    MARKER_CACHE_FILE,
    WHITENESS_CACHE_FILE,
    OMR_MODE_THRESHOLD,
    OMR_MODE_KMEANS,
    KMEANS_N_CLUSTERS,
    KMEANS_MIN_SAMPLES,
    MARK_FORMAT_STANDARD,
    MARK_FORMAT_MULTI_DIGIT,
    MULTI_DIGIT_VALUE_TO_SYMBOL,
    escape_excel_formula,
    get_excel_font_family,
)
from image_alignment import (
    imread_unicode,
    detect_corner_markers,
    apply_perspective_transform,
    compute_output_scale,
)

logger = logging.getLogger(__name__)


def parse_excel_coordinates(excel_path, skip_questions=0):
    """
    Mark2の座標定義Excelファイルをパースして座標リストを取得
    
    Args:
        excel_path: 座標定義Excelファイルのパス
        skip_questions: スキップする問題数（出席番号エリアなど）
                        ※注意: この関数ではスキップ処理を行いません。
                        OMR出力の列ズレ防止のため全座標を読み込みます。
                        実際のスキップ判定はscore_answers() / load_mark2_results()で行います。
    
    Returns:
        coordinates: マーク領域のリスト
        question_groups: 設問ごとのグループ情報
    """
    df = pd.read_excel(excel_path, header=None)

    coordinates = []
    question_groups = {}  # 設問番号 -> 選択肢群の範囲
    renumber_offset = 0  # 再採番用のオフセット

    # row0 の値ヘッダ(base_col=4,8,...に置かれた選択肢の値)を読み取る。
    # raw_choice はこのヘッダ値であり、標準テンプレート(0〜9)ではヘッダ=列の
    # 出現順と一致するが、複数桁モードのテンプレート(-1〜13)では一致しない。
    # ヘッダが欠損・非数値の列は従来互換で出現順インデックスを使う。
    header_row = df.iloc[0] if len(df) > 0 else None
    header_values = {}
    if header_row is not None:
        for raw_choice_idx in range(20):
            base_col = 4 + (raw_choice_idx * 4)
            if base_col >= len(header_row):
                break
            hv = header_row[base_col]
            if pd.notna(hv):
                try:
                    header_values[raw_choice_idx] = int(float(hv))
                except (ValueError, TypeError):
                    pass

    for row_idx in range(3, len(df)):
        row = df.iloc[row_idx]
        original_question_no = row[0]
        
        # 以前はここでskip_questionsに基づいてスキップしていましたが、
        # OMR出力の列ズレを防ぐため、すべての座標を読み込みます。
        # スキップ判定は採点時（score_answers / load_mark2_results）に行います。
        
        question_no = original_question_no
        question_name = row[1] if pd.notna(row[1]) else f"Q{question_no}"
        
        # この設問の選択肢座標を収集
        question_coords = []
        
        # この設問の選択肢座標を一時リストに収集
        temp_coords = []
        
        # 最大20選択肢まで確認（列がある限り）
        for raw_choice_idx in range(20):
            base_col = 4 + (raw_choice_idx * 4)
            
            if base_col + 3 < len(row):
                pos_x = row[base_col]
                pos_y = row[base_col + 1]
                size_x = row[base_col + 2]
                size_y = row[base_col + 3]
                
                if pd.notna(pos_x) and pd.notna(pos_y) and pd.notna(size_x) and pd.notna(size_y):
                    try:
                        coord = {
                            'question_no': question_no,
                            'question': question_name,
                            # choiceは後でX座標順に割り振るため、ここでは仮の値。
                            # raw_choiceは row0 の値ヘッダ(欠損時は出現順)
                            'raw_choice': header_values.get(raw_choice_idx, raw_choice_idx),
                            'x': int(pos_x),
                            'y': int(pos_y),
                            'width': int(size_x),
                            'height': int(size_y)
                        }
                        temp_coords.append(coord)
                    except (ValueError, TypeError):
                        continue
            else:
                break
        
        # X座標でソート（左から右へ）
        temp_coords.sort(key=lambda c: c['x'])
        
        # ソート順にchoice番号（0, 1, 2...）を割り当てて正式なリストに追加
        for i, coord in enumerate(temp_coords):
            coord['choice'] = i
            # raw_choiceは不要なら削除、デバッグ用に残しても良い
            coordinates.append(coord)
            question_coords.append(coord)
        
        # 設問ごとの選択肢群の範囲を計算
        if question_coords:
            x_list = [c['x'] for c in question_coords] + [c['x'] + c['width'] for c in question_coords]
            y_list = [c['y'] for c in question_coords] + [c['y'] + c['height'] for c in question_coords]
            
            min_x = min(x_list)
            max_x = max(x_list)
            min_y = min(y_list)
            max_y = max(y_list)
            
            question_groups[question_no] = {
                'question_name': question_name,
                'choices_bbox': {'x': min_x, 'y': min_y, 'width': max_x - min_x, 'height': max_y - min_y},
                'min_x': min_x,
                'min_y': min_y,
                'height': max_y - min_y
            }
    
    return coordinates, question_groups


def save_template_coordinates_debug(coordinates, output_path):
    """座標リストをCSVファイルに保存（デバッグ用・静的）"""
    try:
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            # ヘッダー
            writer.writerow(['question_no', 'question_name', 'choice', 'x', 'y', 'width', 'height'])
            # データ
            for coord in coordinates:
                writer.writerow([
                    coord['question_no'],
                    coord['question'],
                    coord['choice'],
                    coord['x'],
                    coord['y'],
                    coord['width'],
                    coord['height']
                ])
        logger.info("テンプレート座標データを保存しました: %s", output_path)
    except Exception as e:
        logger.error("テンプレート座標データの保存に失敗しました: %s", e)


def load_coordinates_from_csv(csv_path):
    """CSVファイルから座標リストを読み込む"""
    coordinates = []
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                coord = {
                    'question_no': int(row['question_no']),
                    'question': row['question_name'],
                    'choice': int(row['choice']),
                    'x': int(row['x']),
                    'y': int(row['y']),
                    'width': int(row['width']),
                    'height': int(row['height'])
                }
                coordinates.append(coord)
        return coordinates
    except Exception as e:
        logger.error("座標データの読み込みに失敗しました: %s", e)
        return []


def draw_all_areas(image, coordinates, question_groups):
    """
    全てのエリアを描画
    
    1. マーク領域（選択肢）- 緑の枠
    2. 設問単位の選択肢群 - 赤の枠
    """
    result_image = image.copy()
    
    # 1. マーク領域を描画（緑の枠）
    mark_count = 0
    for coord in coordinates:
        x = coord['x']
        y = coord['y']
        w = coord['width']
        h = coord['height']
        
        cv2.rectangle(result_image, (x, y), (x + w, y + h), (0, 255, 0), 2)
        mark_count += 1
    
    # 2. 設問ごとの選択肢群を描画（赤の枠）
    group_count = 0
    for question_no, group_data in question_groups.items():
        choices_bbox = group_data['choices_bbox']
        cv2.rectangle(result_image,
                     (choices_bbox['x'], choices_bbox['y']),
                     (choices_bbox['x'] + choices_bbox['width'], choices_bbox['y'] + choices_bbox['height']),
                     (0, 0, 255), 2)
        group_count += 1
        
    return result_image, mark_count, group_count


def generate_template(coord_excel_path, output_folder, skip_questions=0):
    """
    採点用正答データExcelを生成
    
    Args:
        coord_excel_path: 座標定義ファイルのパス
        output_folder: 出力先フォルダ
        skip_questions: スキップする問題数（出席番号エリアなど）
    
    Returns:
        template_path: 生成されたテンプレートファイルのパス
    """
    # 座標ファイルを読み込み
    df_coord = pd.read_excel(coord_excel_path, header=None)
    
    template_data = []
    
    # 3行目から問題データを読み取り
    for row_idx in range(3, len(df_coord)):
        row = df_coord.iloc[row_idx]
        original_question_no = row[0]
        
        if pd.isna(original_question_no):
            continue
        
        # スキップする問題は除外
        if original_question_no <= skip_questions:
            continue
        
        # 再採番された問題番号
        question_no = original_question_no - skip_questions
        
        template_data.append({
            '問題番号': question_no,
            '正答': '',    # 空欄（ユーザーが入力）
            '配点': '',    # 空欄（ユーザーが入力）
            '観点': '',    # 空欄（ユーザーが入力）
            '特例': '',    # 任意入力（ドロップダウン: 全員正解。不適切問題の救済措置）
            '問題概要': ''  # 任意入力（20字程度。CTT/R連携レポートに表示される）
        })
    
    # DataFrameに変換
    df_template = pd.DataFrame(template_data)

    # Excelファイルとして出力
    output_folder = Path(output_folder)
    output_folder.mkdir(exist_ok=True)
    template_path = output_folder / ANSWER_KEY_FILE

    # ⚠ 既存のanswer_key.xlsxがある場合は上書きしない（ユーザー入力済みの正答・配点を保護）
    if template_path.exists():
        logger.info("テンプレートが既に存在します。上書きしません: %s", template_path.name)
        return template_path

    _write_styled_template(df_template, template_path)

    return template_path


def _write_styled_template(df_template, template_path):
    """answer_key.xlsx を装飾付きで書き出す。

    ユーザーがExcelで直接開いて正答・配点・観点を手入力するファイルのため、
    summary_generator.py と同じハウススタイル(青ヘッダー・罫線・交互背景色・
    ヘッダー行固定)を適用して視認性を上げる。
    列構成・行位置は従来の to_excel 出力と同一(1行目ヘッダー、2行目以降データ)。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    header_font = Font(name=get_excel_font_family(), bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    data_font = Font(name=get_excel_font_family(), size=10)
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF'),
    )
    light_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
    alt_fill = PatternFill(start_color='E9EFF5', end_color='E9EFF5', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')

    wb = Workbook()
    ws = wb.active

    columns = list(df_template.columns)
    for ci, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    for ri, (_, row) in enumerate(df_template.iterrows(), 2):
        fill = light_fill if ri % 2 == 0 else alt_fill
        for ci, col_name in enumerate(columns, 1):
            value = row[col_name]
            if value == '':
                value = None  # 空欄は空セルとして出力(従来のto_excelと同じ読み込み挙動)
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font = data_font
            cell.fill = fill
            cell.border = thin_border
            if col_name in ('問題番号', '配点', '観点', '特例'):
                cell.alignment = center

    column_widths = {'問題番号': 10, '正答': 12, '配点': 8, '観点': 8, '特例': 12, '問題概要': 30}
    for ci, col_name in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = column_widths.get(col_name, 12)

    # 特例列にドロップダウン（空欄 or 全員正解）を設定する。
    # 不適切問題の救済措置として使う想定のため、自由入力は許可しない
    if '特例' in columns and len(df_template) > 0:
        special_col_letter = get_column_letter(columns.index('特例') + 1)
        dv = DataValidation(type='list', formula1='"全員正解"', allow_blank=True)
        dv.error = '「全員正解」または空欄のみ入力できます'
        dv.errorTitle = '特例区分'
        dv.prompt = '不適切問題を全員正解として救済する場合に選択（配点は必須のまま。正答は空欄でも可）'
        dv.promptTitle = '特例区分'
        ws.add_data_validation(dv)
        dv.add(f'{special_col_letter}2:{special_col_letter}{len(df_template) + 1}')

    ws.freeze_panes = 'A2'
    wb.save(str(template_path))


def save_coordinates_to_csv(csv_path, all_data):
    """
    座標データをCSV形式で保存（画像ごとの補正後座標）
    
    CSV形式:
    image_path,question_no,choices_bbox,mark_coords
    
    choices_bbox: x;y;w;h
    mark_coords: choice0_x;y;w;h|choice1_x;y;w;h|...
    
    Raises:
        PermissionError: ファイルが他のアプリで開かれている場合
    """
    try:
        _save_coordinates_to_csv_impl(csv_path, all_data)
    except PermissionError:
        logger.warning("%s への保存に失敗しました。ファイルが別のアプリで開かれている可能性があります。", csv_path)
        raise


def _save_coordinates_to_csv_impl(csv_path, all_data):
    """save_coordinates_to_csvの実装部"""
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        
        # ヘッダー
        writer.writerow([
            'image_path',
            'question_no',
            'choices_bbox',
            'mark_coords'
        ])
        
        # データ行
        for data in all_data:
            image_path = data['image_path']
            question_no = data['question_no']
            
            # choices_bbox形式: x;y;w;h
            choices_bbox = data['choices_bbox']
            choices_str = f"{choices_bbox['x']};{choices_bbox['y']};{choices_bbox['width']};{choices_bbox['height']}"
            
            # マーク座標: choice0_x;y;w;h|choice1_x;y;w;h|...
            mark_list = []
            for mark in data['mark_coords']:
                mark_list.append(f"{mark['x']};{mark['y']};{mark['width']};{mark['height']}")
            mark_str = '|'.join(mark_list)
            
            writer.writerow([
                image_path,
                question_no,
                choices_str,
                mark_str
            ])


def recognize_marks(image, coordinates, color_threshold=0.1, area_threshold=0.4):
    """
    マーク認識を行う (Mark2OSSロジック準拠)
    
    Args:
        image: 補正後の画像 (Gray or BGR)
        coordinates: マーク領域のリスト
        color_threshold: 画素値の閾値 (0.0-1.0). (1 - color_threshold) * 255 より暗い画素をマークとみなす.
                         Default: 0.1 (255 * 0.9 = 229.5未満をマークとする)
        area_threshold: 面積閾値 (0.0-1.0). マーク画素の割合がこれを超えたらマークとみなす.
                        Default: 0.4 (40%以上)
    
    Returns:
        results: {question_no: [choice_idx, ...]}
    """
    if len(image.shape) == 3:
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    else:
        gray = image
        
    results = {}
    
    # 閾値計算
    pixel_threshold = int((1.0 - color_threshold) * 255)
    
    # 二値化
    # 画素値 < pixel_threshold (暗い) -> 255 (白/カウント対象)
    _, binary = cv2.threshold(gray, pixel_threshold, 255, cv2.THRESH_BINARY_INV)
    
    for coord in coordinates:
        q_no = coord['question_no']
        choice_idx = coord['choice']
        x, y, w, h = coord['x'], coord['y'], coord['width'], coord['height']
        
        # ROI抽出
        roi = binary[y:y+h, x:x+w]
        
        # マーク画素数をカウント
        marked_pixels = cv2.countNonZero(roi)
        total_pixels = w * h
        
        if total_pixels == 0:
            continue
            
        ratio = marked_pixels / total_pixels
        
        if ratio > area_threshold:
            if q_no not in results:
                results[q_no] = []
            results[q_no].append(choice_idx)
            
    return results


# ========================================
# K-means クラスタリング方式 (v4.5)
# ========================================

def extract_mark_features(gray_image, coordinates):
    """
    マーク領域ごとに5次元ローカル特徴量を抽出する。

    Args:
        gray_image: グレースケール画像 (射影変換済み)
        coordinates: マーク領域座標リスト

    Returns:
        features: np.ndarray (N, 5) — [filled_ratio, mean_inv_brightness,
                  dark_pixel_ratio, std_inv_brightness, center_edge_ratio]
        meta: list[dict] — 各要素の question_no, choice 情報
    """
    features = []
    meta = []
    for coord in coordinates:
        x, y, w, h = coord['x'], coord['y'], coord['width'], coord['height']
        roi = gray_image[y:y+h, x:x+w]
        total_pixels = w * h
        if total_pixels == 0:
            features.append([0.0, 0.0, 0.0, 0.0, 0.0])
            meta.append({'question_no': coord['question_no'], 'choice': coord['choice']})
            continue

        # 二値化 (Otsu)
        _, binary = cv2.threshold(roi, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        filled_ratio = cv2.countNonZero(binary) / total_pixels

        # 反転輝度 (0=白, 255=黒)
        inv = 255 - roi.astype(np.float32)
        mean_inv = float(np.mean(inv)) / 255.0
        std_inv = float(np.std(inv)) / 255.0

        # 暗画素率 (値<=128 の画素)
        dark_pixel_ratio = float(np.sum(roi <= 128)) / total_pixels

        # 中心/辺縁濃度比 — マークの塗り方パターンを捉える
        rh, rw = roi.shape
        if rh >= 4 and rw >= 4:
            mh, mw = max(1, rh // 4), max(1, rw // 4)
            center = roi[mh:rh - mh, mw:rw - mw]
            top = roi[:mh, :]
            bot = roi[rh - mh:, :]
            left = roi[:, :mw]
            right = roi[:, rw - mw:]
            edge = np.concatenate([top.ravel(), bot.ravel(), left.ravel(), right.ravel()])
            center_dark = float(255 - np.mean(center)) / 255.0
            edge_dark = float(255 - np.mean(edge)) / 255.0
            center_edge_ratio = center_dark / max(edge_dark, 0.001)
        else:
            center_edge_ratio = 1.0

        features.append([filled_ratio, mean_inv, dark_pixel_ratio, std_inv, center_edge_ratio])
        meta.append({'question_no': coord['question_no'], 'choice': coord['choice']})

    return np.array(features, dtype=np.float64), meta


def recognize_marks_kmeans(image, coordinates, n_clusters=KMEANS_N_CLUSTERS,
                           min_samples=KMEANS_MIN_SAMPLES,
                           fallback_area_threshold=0.4,
                           fallback_color_threshold=0.1):
    """
    K-means クラスタリングによるマーク認識 (v4.5 enriched)。

    5次元ローカル特徴量に加え、シート内正規化・設問内コントラストの
    コンテキスト特徴量を追加した7次元空間で K-means(K=2) を実行する。

    特徴量 (7次元):
      - filled_ratio, mean_inv, dark_pixel_ratio, std_inv: ローカル特徴量
      - center_edge_ratio: 中心/辺縁の濃度比 (塗り方パターン)
      - normalized_filled: シート内最大 filled_ratio で正規化 (生徒の癖を吸収)
      - question_contrast: 設問内での相対的な濃さ差分 (消しゴム痕を識別)

    Args:
        image: 補正後の画像 (Gray or BGR)
        coordinates: マーク領域のリスト
        n_clusters: クラスタ数 (デフォルト 2)
        min_samples: 最小サンプル数 (デフォルト 50)
        fallback_area_threshold: フォールバック時の面積閾値
        fallback_color_threshold: フォールバック時の色閾値

    Returns:
        results: {question_no: [choice_idx, ...]}
        kmeans_info: dict — K-means の詳細情報 (レポート生成用)。
                     フォールバック時は None。
    """
    if len(image.shape) == 3:
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    else:
        gray = image

    if len(coordinates) < min_samples:
        logger.info("K-means: サンプル数 %d < %d、閾値方式にフォールバック",
                     len(coordinates), min_samples)
        results = recognize_marks(image, coordinates,
                                  color_threshold=fallback_color_threshold,
                                  area_threshold=fallback_area_threshold)
        return results, None

    # --- ローカル特徴量抽出 (N, 5) ---
    local_features, meta = extract_mark_features(gray, coordinates)

    # --- コンテキスト特徴量の計算 ---
    filled_col = local_features[:, 0]  # filled_ratio
    student_max_filled = float(filled_col.max()) if len(filled_col) > 0 else 0.001

    # normalized_filled: シート内の最大 filled_ratio で正規化
    normalized_filled = filled_col / max(student_max_filled, 0.001)

    # question_contrast: 設問内での相対的な差分
    from collections import defaultdict
    q_groups = defaultdict(list)
    for i, m in enumerate(meta):
        q_groups[m['question_no']].append((i, filled_col[i]))

    question_contrast = np.zeros(len(meta), dtype=np.float64)
    for q_no, members in q_groups.items():
        sorted_members = sorted(members, key=lambda x: x[1], reverse=True)
        for rank, (idx, val) in enumerate(sorted_members):
            if rank == 0 and len(sorted_members) >= 2:
                # 最も濃い → 2番目との差
                question_contrast[idx] = val - sorted_members[1][1]
            elif rank == 0:
                question_contrast[idx] = 0.0
            else:
                # 2番目以降 → 最大との差 (負の値)
                question_contrast[idx] = val - sorted_members[0][1]

    # --- 7次元特徴量行列 ---
    features = np.column_stack([
        local_features,           # (N, 5): filled, mean_inv, dark_pixel, std_inv, center_edge
        normalized_filled,        # (N,)
        question_contrast,        # (N,)
    ])  # → (N, 7)

    try:
        from sklearn.preprocessing import StandardScaler
        from sklearn.cluster import KMeans
    except ImportError as _e:
        raise RuntimeError(
            "K-meansクラスタリングに必要な scikit-learn が見つかりません。\n"
            "pip install scikit-learn を実行してください。\n"
            f"（詳細: {_e}）"
        ) from _e

    scaler = StandardScaler()
    X = scaler.fit_transform(features)

    kmeans = KMeans(n_clusters=n_clusters, n_init=10, random_state=42)
    labels = kmeans.fit_predict(X)

    # filled_ratio (index=0) のクラスタ平均が高い方を「マーク済み」と判定
    cluster_means = []
    for c in range(n_clusters):
        mask = labels == c
        if mask.sum() > 0:
            cluster_means.append(float(features[mask, 0].mean()))
        else:
            cluster_means.append(0.0)
    marked_cluster = int(np.argmax(cluster_means))

    # filled_ratio の境界値（2クラスタの中間点）
    sorted_clusters = sorted(range(n_clusters), key=lambda c: cluster_means[c])
    if len(sorted_clusters) >= 2:
        c_low, c_high = sorted_clusters[0], sorted_clusters[1]
        cutoff = (cluster_means[c_low] + cluster_means[c_high]) / 2.0
    else:
        cutoff = 0.5

    results = {}
    for i, m in enumerate(meta):
        if labels[i] == marked_cluster:
            q_no = m['question_no']
            if q_no not in results:
                results[q_no] = []
            results[q_no].append(m['choice'])

    # --- 信頼度スコア (設問単位) ---
    # cutoff からの距離が小さい設問 = 判定に不確実性がある
    question_confidences = {}
    for q_no, members in q_groups.items():
        min_margin = float('inf')
        for idx, val in members:
            margin = abs(val - cutoff)
            if margin < min_margin:
                min_margin = margin
        question_confidences[q_no] = round(min_margin, 4)

    # ROI サムネイル (16×16 → base64) — レポートのホバー表示用
    import base64 as _b64
    roi_thumbnails = []
    for coord in coordinates:
        x, y, w, h = coord['x'], coord['y'], coord['width'], coord['height']
        roi = gray[y:y+h, x:x+w]
        if roi.size > 0:
            roi_small = cv2.resize(roi, (16, 16))
        else:
            roi_small = np.full((16, 16), 200, dtype=np.uint8)
        _, buf = cv2.imencode(".png", roi_small)
        roi_thumbnails.append(_b64.b64encode(buf).decode("utf-8"))

    kmeans_info = {
        'features': features,
        'labels': labels,
        'meta': meta,
        'cluster_means': cluster_means,
        'marked_cluster': marked_cluster,
        'cutoff': cutoff,
        'n_marked': int((labels == marked_cluster).sum()),
        'n_empty': int((labels != marked_cluster).sum()),
        'scaler_mean': scaler.mean_.tolist(),
        'scaler_scale': scaler.scale_.tolist(),
        'roi_thumbnails': roi_thumbnails,
        'question_confidences': question_confidences,
        'student_max_filled': student_max_filled,
    }

    return results, kmeans_info


def save_recognition_results(output_path, recognition_results, all_questions, question_names=None, choice_counts=None, coordinates=None, mark_format=MARK_FORMAT_STANDARD):
    """
    認識結果をExcelファイルに保存 (Mark2OSS Survey.cs準拠)
    スタイリング: ヘッダー装飾, 罫線, NoMark背景色(オレンジ), DoubleMark背景色(薄い赤), ウィンドウ枠固定

    Args:
        output_path: 出力Excelファイルパス
        recognition_results: 認識結果リスト
        all_questions: 全設問番号リスト
        question_names: 設問名辞書 (optional)
        choice_counts: 設問番号 -> 選択肢数 の辞書 (optional, 未指定時は10)
        mark_format: MARK_FORMAT_MULTI_DIGIT ならセル値を紙面記号
                     (-1→「-」、10〜13→a〜d)に変換して出力する
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # --- スタイル定義 ---
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    # raw_choiceルックアップ: {q_no: {sorted_choice_idx: raw_choice_value}}
    # Excelの列ヘッダ値（raw_choice）を表示値として使用する。
    # 従来の (c+1)%num_choices は横並び10列専用だったが、
    # raw_choice は縦並び・異なる選択肢数のテンプレートでも正しい値を返す。
    choice_to_display = {}
    if coordinates:
        for c in coordinates:
            q = c['question_no']
            if q not in choice_to_display:
                choice_to_display[q] = {}
            choice_to_display[q][c['choice']] = c['raw_choice']

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    label_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    label_font = Font(bold=True, size=9)
    no_mark_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    double_mark_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')

    # --- Row 1: ヘッダー (No, File, 設問番号) ---
    header_values = ['No', 'File'] + [str(q) for q in all_questions]
    ws.append(header_values)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # --- Row 2: 設問名ラベル ---
    label_values = ['', '']
    if question_names:
        for q in all_questions:
            label_values.append(question_names.get(q, ''))
    else:
        label_values.extend([''] * len(all_questions))
    ws.append(label_values)
    for cell in ws[2]:
        cell.font = label_font
        cell.fill = label_fill
        cell.alignment = center_align
        cell.border = thin_border

    # --- データ行 ---
    for idx, res in enumerate(recognition_results):
        row_values = [idx + 1, escape_excel_formula(res['image'])]
        marks = res['marks']

        for q_no in all_questions:
            if q_no in marks:
                choices = marks[q_no]
                num_choices = choice_counts.get(q_no, 10) if choice_counts else 10
                val_strs = []
                for c in sorted(choices):
                    # raw_choice（Excelの列ヘッダ値）を表示値として使用
                    # 座標パース時に保存された raw_choice は、テンプレートの
                    # レイアウト（横並び・縦並び）に依存しない正しい値を持つ。
                    if q_no in choice_to_display and c in choice_to_display[q_no]:
                        val = choice_to_display[q_no][c]
                    else:
                        # フォールバック（coordinates未指定時の後方互換）
                        val = (c + 1) % num_choices
                    if mark_format == MARK_FORMAT_MULTI_DIGIT:
                        # 複数桁モード: ヘッダ値を紙面記号(-, 0〜9, a〜d)へ変換
                        val = MULTI_DIGIT_VALUE_TO_SYMBOL.get(val, str(val))
                    val_strs.append(str(val))
                row_values.append(';'.join(val_strs))
            else:
                row_values.append('')  # No mark

        ws.append(row_values)
        row_idx = ws.max_row

        for col_idx in range(1, len(row_values) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = center_align
            elif col_idx >= 3:
                cell.alignment = center_align
                cell_value = cell.value
                if cell_value is None or (isinstance(cell_value, str) and cell_value == ''):
                    cell.fill = no_mark_fill
                elif isinstance(cell_value, str) and ';' in cell_value:
                    cell.fill = double_mark_fill

    # --- 列幅 ---
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 28
    for col_idx in range(3, len(all_questions) + 3):
        col_letter = ws.cell(1, col_idx).column_letter
        ws.column_dimensions[col_letter].width = 8

    # --- ウィンドウ枠固定 (ヘッダー2行 + No/File列) ---
    ws.freeze_panes = 'C3'

    try:
        wb.save(output_path)
    except PermissionError:
        logger.warning("%s への保存に失敗しました。ファイルが別のアプリで開かれている可能性があります。", output_path)
        raise


def _process_single_image(args: tuple) -> dict:
    """
    1枚の画像を処理するワーカー関数（ProcessPoolExecutor 用）。

    モジュールレベルに定義することで pickle 可能。
    各ワーカープロセスで独立して実行される。

    Args:
        args: (image_path_str, boxed_folder_str, clean_folder_str, coordinates,
               question_groups, color_threshold, area_threshold, omr_mode)

    Returns:
        dict with keys: filename, marks, marker_data, csv_data, success, kmeans_info
    """
    # 後方互換: 8要素なら omr_mode あり、7要素なら閾値方式
    if len(args) == 8:
        (image_path_str, boxed_folder_str, clean_folder_str, coordinates,
         question_groups, color_threshold, area_threshold, omr_mode) = args
    else:
        (image_path_str, boxed_folder_str, clean_folder_str, coordinates,
         question_groups, color_threshold, area_threshold) = args
        omr_mode = OMR_MODE_THRESHOLD

    image_path = Path(image_path_str)
    boxed_folder = Path(boxed_folder_str)
    clean_folder = Path(clean_folder_str) if clean_folder_str else None

    with open(str(image_path), 'rb') as f:
        image_data_bytes = f.read()
    image = cv2.imdecode(np.frombuffer(image_data_bytes, np.uint8), cv2.IMREAD_COLOR)

    if image is None:
        raise ValueError(f"画像を読み込めません: {image_path.name}")

    markers = detect_corner_markers(image, debug=False)
    corrected_image, _ = apply_perspective_transform(image, markers)
    marker_data = [
        pt.tolist() if hasattr(pt, 'tolist') else list(pt) for pt in markers
    ]

    # OMR認識
    kmeans_info = None
    if omr_mode == OMR_MODE_KMEANS:
        marks, kmeans_info = recognize_marks_kmeans(
            corrected_image, coordinates,
            fallback_area_threshold=area_threshold,
            fallback_color_threshold=color_threshold,
        )
    else:
        marks = recognize_marks(
            corrected_image, coordinates,
            color_threshold=color_threshold, area_threshold=area_threshold,
        )

    # --- 白さキャッシュ用: 設問ごとの平均輝度を計算 ---
    whiteness = {}
    try:
        _gray_for_wh = cv2.cvtColor(corrected_image, cv2.COLOR_BGR2GRAY) if len(corrected_image.shape) == 3 else corrected_image
        if kmeans_info is not None:
            _wh_features = kmeans_info['features']   # (N, 7)
            _wh_meta = kmeans_info['meta']
        else:
            _wh_features, _wh_meta = extract_mark_features(_gray_for_wh, coordinates)  # (N, 5)
        # 設問ごとに mean_inv (index=1) の平均から輝度を算出
        _q_brightness = {}
        _q_counts = {}
        for i, m in enumerate(_wh_meta):
            q_no = m['question_no']
            brightness = (1.0 - float(_wh_features[i, 1])) * 255.0
            if q_no not in _q_brightness:
                _q_brightness[q_no] = 0.0
                _q_counts[q_no] = 0
            _q_brightness[q_no] += brightness
            _q_counts[q_no] += 1
        whiteness = {
            str(int(q)): round(_q_brightness[q] / _q_counts[q], 2)
            for q in _q_brightness if _q_counts[q] > 0
        }
    except Exception as e:
        logger.warning("白さ計算に失敗しました（MarkCheckerでフォールバック計算されます）: %s", e)

    # 枠描画
    result_image, _mark_count, _group_count = draw_all_areas(
        corrected_image, coordinates, question_groups,
    )

    # 認識結果をオーバーレイ描画
    for q_no, choices in marks.items():
        for c_idx in choices:
            target = next(
                (c for c in coordinates
                 if c['question_no'] == q_no and c['choice'] == c_idx),
                None,
            )
            if target:
                cv2.rectangle(
                    result_image,
                    (target['x'], target['y']),
                    (target['x'] + target['width'], target['y'] + target['height']),
                    (255, 0, 0), 2,
                )

    # boxed画像を保存
    output_path = boxed_folder / image_path.name
    _, encoded = cv2.imencode('.jpg', result_image)
    with open(str(output_path), 'wb') as f:
        f.write(encoded)

    # クリーン画像（枠描画なし）を保存 — 記述式採点プレビュー用
    if clean_folder is not None:
        clean_output_path = clean_folder / image_path.name
        _, clean_encoded = cv2.imencode('.jpg', corrected_image)
        with open(str(clean_output_path), 'wb') as f:
            f.write(clean_encoded)

    # CSVデータ構築
    csv_data = []
    for question_no, group_data in question_groups.items():
        question_marks = [c for c in coordinates if c['question_no'] == question_no]
        csv_data.append({
            'image_path': str(image_path.name),
            'question_no': question_no,
            'choices_bbox': group_data['choices_bbox'],
            'mark_coords': question_marks,
        })

    return {
        'filename': image_path.name,
        'marks': marks,
        'marker_data': marker_data,
        'csv_data': csv_data,
        'kmeans_info': kmeans_info,
        'whiteness': whiteness,
        'success': True,
    }


def generate_kmeans_report(output_path, all_kmeans_infos):
    """
    K-means クラスタリング結果の HTML レポートを生成する。

    全画像の特徴量を集約し、filled_ratio ヒストグラム・PCA 散布図・
    クラスタ統計をインタラクティブに表示する。
    PCA 散布図ではマウスホバーで個別マーク領域の画像を表示する。

    Args:
        output_path: 出力 HTML ファイルパス
        all_kmeans_infos: list[dict] — 各要素に 'filename', 'info' キー
    """
    # 全画像の特徴量・ラベル・メタ・ROI を集約
    # 重要: 各画像の K-means は独立して実行されるため、クラスタラベル（0/1）の
    # 割り当ては画像ごとに異なる。レポートではラベルを統一して
    # 「1 = マーク済み, 0 = 空白」に正規化する。
    all_features = []
    all_labels = []
    all_filenames = []
    all_meta = []
    all_rois = []
    for entry in all_kmeans_infos:
        info = entry['info']
        n = len(info['labels'])
        all_features.append(info['features'])

        # ラベル正規化: marked_cluster → 常に 1, empty → 常に 0
        raw_labels = info['labels'].copy()
        mc = info['marked_cluster']
        if mc == 0:
            # この画像ではクラスタ0がマーク済み → 反転 (0→1, 1→0)
            aligned_labels = 1 - raw_labels
        else:
            aligned_labels = raw_labels
        all_labels.append(aligned_labels)

        all_filenames.extend([entry['filename']] * n)
        all_meta.extend(info.get('meta', [{'question_no': 0, 'choice': 0}] * n))
        rois = info.get('roi_thumbnails', [''] * n)
        all_rois.extend(rois)

    features = np.vstack(all_features)
    labels = np.concatenate(all_labels)
    total = len(labels)

    # 統計サマリー（正規化後は marked_cluster = 1 で統一）
    marked_cluster = 1
    # カットオフは全画像の平均を使用
    cutoff = float(np.mean([e['info']['cutoff'] for e in all_kmeans_infos]))
    n_marked = int((labels == marked_cluster).sum())
    n_empty = total - n_marked

    # filled_ratio ヒストグラムデータ
    filled = features[:, 0]
    hist_marked = filled[labels == marked_cluster].tolist()
    hist_empty = filled[labels != marked_cluster].tolist()

    # PCA 2D (全データ)
    from sklearn.decomposition import PCA
    from sklearn.preprocessing import StandardScaler
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(features)
    pca = PCA(n_components=2)
    X_pca = pca.fit_transform(X_scaled)
    variance_ratio = pca.explained_variance_ratio_

    # PCA データを JSON 用にサンプリング（最大 5000 点）
    max_points = 5000
    if total > max_points:
        rng = np.random.RandomState(42)
        idx = rng.choice(total, max_points, replace=False)
    else:
        idx = np.arange(total)

    # 散布図データ（ホバー用メタ情報込み）
    scatter_points = []
    for i in idx:
        scatter_points.append({
            'x': round(float(X_pca[i, 0]), 4),
            'y': round(float(X_pca[i, 1]), 4),
            'lbl': int(labels[i]),
            'filled': round(float(features[i, 0]), 4),
            'img': all_filenames[i],
            'q': int(all_meta[i]['question_no']),
            'ch': int(all_meta[i]['choice']),
            'roi': all_rois[i] if i < len(all_rois) else '',
        })

    # クラスタごとにデータセットを分割
    empty_pts = [p for p in scatter_points if p['lbl'] != marked_cluster]
    marked_pts = [p for p in scatter_points if p['lbl'] == marked_cluster]
    # 境界クラスタ判定（3クラスタ以上の将来拡張用）
    boundary_pts: list[dict] = []

    html = f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="utf-8">
<title>K-means OMR Report</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<style>
body {{ font-family: 'Yu Gothic UI', sans-serif; margin: 20px; background: #f5f5f5; }}
.container {{ max-width: 1200px; margin: auto; }}
h1 {{ color: #1976D2; }}
h3 {{ color: #333; }}
.stats {{ display: flex; gap: 20px; margin: 20px 0; }}
.stat-card {{ background: white; border-radius: 8px; padding: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); flex: 1; text-align: center; }}
.stat-card .value {{ font-size: 2em; font-weight: bold; color: #1976D2; }}
.stat-card .label {{ color: #666; margin-top: 5px; }}
.chart-row {{ display: flex; gap: 20px; margin: 20px 0; }}
.chart-box {{ background: white; border-radius: 8px; padding: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); flex: 1; }}
canvas {{ max-width: 100%; }}
.pca-section {{ background: white; border-radius: 8px; padding: 20px; margin: 20px 0; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
.pca-desc {{ color: #666; font-size: 0.9em; margin-bottom: 10px; }}
.pca-legend {{ display: flex; gap: 16px; justify-content: center; margin-bottom: 10px; font-size: 0.9em; }}
.pca-legend span {{ display: inline-flex; align-items: center; gap: 4px; }}
.pca-legend .dot {{ width: 10px; height: 10px; border-radius: 50%; display: inline-block; }}
#hover-info {{ margin-top: 8px; font-size: 12px; color: #555; min-height: 24px; padding: 4px 8px; border-top: 1px solid #eee; }}
</style>
</head>
<body>
<div class="container">
<h1>K-means OMR Report</h1>
<p>処理画像数: {len(all_kmeans_infos)} | 総領域数: {total:,}</p>

<div class="stats">
  <div class="stat-card">
    <div class="value">{n_marked:,}</div>
    <div class="label">マーク済み ({n_marked/total*100:.1f}%)</div>
  </div>
  <div class="stat-card">
    <div class="value">{n_empty:,}</div>
    <div class="label">空白 ({n_empty/total*100:.1f}%)</div>
  </div>
  <div class="stat-card">
    <div class="value">{cutoff:.4f}</div>
    <div class="label">filled_ratio カットオフ</div>
  </div>
</div>

<div class="chart-row">
  <div class="chart-box">
    <h3>filled_ratio 分布</h3>
    <canvas id="histChart"></canvas>
  </div>
</div>

<div class="pca-section">
  <h3>K-means 散布図 (PCA 2次元投影)</h3>
  <p class="pca-desc">7次元特徴空間をPCAで2次元に圧縮。各点にホバーするとマーク領域の詳細が表示されます。</p>
  <div class="pca-legend">
    <span><span class="dot" style="background:#4CAF50"></span> 未マーク (Empty)</span>
    <span><span class="dot" style="background:#EF5350"></span> マーク済み (Marked)</span>
  </div>
  <canvas id="pcaChart" height="400"></canvas>
  <div id="hover-info"></div>
</div>

</div>

<script>
const histMarked = {json.dumps(hist_marked[:2000])};
const histEmpty = {json.dumps(hist_empty[:2000])};
const cutoff = {cutoff};

// Histogram
function buildHistogram(data, bins, min, max) {{
  const step = (max - min) / bins;
  const counts = new Array(bins).fill(0);
  const labels = [];
  for (let i = 0; i < bins; i++) {{
    labels.push((min + step * (i + 0.5)).toFixed(3));
    for (const v of data) {{
      if (v >= min + step * i && v < min + step * (i + 1)) counts[i]++;
    }}
  }}
  return {{ labels, counts }};
}}

const hm = buildHistogram(histMarked, 50, 0, 1);
const he = buildHistogram(histEmpty, 50, 0, 1);

new Chart(document.getElementById('histChart'), {{
  type: 'bar',
  data: {{
    labels: hm.labels,
    datasets: [
      {{ label: 'Empty', data: he.counts, backgroundColor: 'rgba(66,133,244,0.5)' }},
      {{ label: 'Marked', data: hm.counts, backgroundColor: 'rgba(244,67,54,0.5)' }}
    ]
  }},
  options: {{
    scales: {{ x: {{ title: {{ display: true, text: 'filled_ratio' }} }},
               y: {{ title: {{ display: true, text: 'Count' }} }} }},
    plugins: {{ legend: {{ position: 'top' }} }}
  }},
  plugins: [{{
    afterDraw(chart) {{
      const xAxis = chart.scales.x;
      const yAxis = chart.scales.y;
      const idx = Math.round(cutoff * 50);
      if (idx >= 0 && idx < xAxis.ticks.length) {{
        const x = xAxis.getPixelForValue(idx);
        const ctx = chart.ctx;
        ctx.save();
        ctx.strokeStyle = '#FF9800';
        ctx.lineWidth = 2;
        ctx.setLineDash([6, 4]);
        ctx.beginPath();
        ctx.moveTo(x, yAxis.top);
        ctx.lineTo(x, yAxis.bottom);
        ctx.stroke();
        ctx.restore();
      }}
    }}
  }}]
}});

// PCA Scatter with hover
const emptyPts = {json.dumps(empty_pts)};
const markedPts = {json.dumps(marked_pts)};
const markedCluster = {marked_cluster};

const chart = new Chart(document.getElementById('pcaChart'), {{
  type: 'scatter',
  data: {{
    datasets: [
      {{
        label: '\u672a\u30de\u30fc\u30af',
        data: emptyPts.map(p => ({{x: p.x, y: p.y}})),
        meta_: emptyPts,
        backgroundColor: 'rgba(76,175,80,0.3)',
        pointRadius: 2,
      }},
      {{
        label: '\u30de\u30fc\u30af\u6e08\u307f',
        data: markedPts.map(p => ({{x: p.x, y: p.y}})),
        meta_: markedPts,
        backgroundColor: 'rgba(239,83,80,0.3)',
        pointRadius: 2,
      }}
    ]
  }},
  options: {{
    responsive: true,
    scales: {{
      x: {{ title: {{ display: true, text: 'PC1 ({variance_ratio[0]*100:.1f}%)' }} }},
      y: {{ title: {{ display: true, text: 'PC2 ({variance_ratio[1]*100:.1f}%)' }} }}
    }},
    plugins: {{ legend: {{ position: 'top' }} }},
    onHover: (event, elements) => {{
      if (elements.length > 0) {{
        const el = elements[0];
        const ds = chart.data.datasets[el.datasetIndex];
        const pt = ds.meta_[el.index];
        const roiHtml = pt.roi
          ? '<img src="data:image/png;base64,' + pt.roi + '" style="width:32px;height:32px;image-rendering:pixelated;vertical-align:middle;margin-left:8px">'
          : '';
        document.getElementById('hover-info').innerHTML =
          '<strong>' + pt.img + '</strong> \u8a2d\u554f' + pt.q + ' \u9078\u629e\u80a2' + (pt.ch+1) + ' | filled=' + pt.filled + ' | role=' + (pt.lbl === markedCluster ? 'marked' : 'empty') + roiHtml;
      }}
    }}
  }}
}});
</script>
</body>
</html>"""

    with open(str(output_path), 'w', encoding='utf-8') as f:
        f.write(html)


def process_box_drawer(image_folder, coord_excel_path, skip_questions=0, output_base_folder=None, debug=False, color_threshold=0.1, area_threshold=0.4, progress_callback=None, cancel_event=None, omr_mode=OMR_MODE_THRESHOLD, mark_format=MARK_FORMAT_STANDARD):
    """
    フォルダ内の画像を一括処理（枠描画 + OMR認識）

    Args:
        progress_callback: 進捗コールバック(current, total)（オプション、GUIプログレスバー用）
        cancel_event: threading.Event — set()されると処理を中断
        omr_mode: OMR認識モード ('threshold' or 'kmeans')
        mark_format: マーク形式。MARK_FORMAT_MULTI_DIGIT なら読取結果Excelの
                     セル値を紙面記号(-, 0〜9, a〜d)で出力する
    """
    start_time = time.time()
    
    image_folder = Path(image_folder)
    coord_excel_path = Path(coord_excel_path)
    
    if output_base_folder is None:
        output_base_folder = image_folder
    else:
        output_base_folder = Path(output_base_folder)
    
    results_folder = output_base_folder / RESULTS_FOLDER
    results_folder.mkdir(exist_ok=True)
    
    boxed_folder = results_folder / BOXED_FOLDER
    boxed_folder.mkdir(exist_ok=True)

    clean_folder = results_folder / CLEAN_FOLDER
    clean_folder.mkdir(exist_ok=True)

    results_data_folder = results_folder / RESULTS_DATA_FOLDER
    results_data_folder.mkdir(exist_ok=True)

    reading_results_folder = results_data_folder / READING_RESULTS_FOLDER_NAME
    reading_results_folder.mkdir(exist_ok=True)
    
    logger.info("出力フォルダ: %s", results_folder)
    logger.info("枠描画結果: %s/", boxed_folder.name)
    logger.info("読取結果: %s/%s/", results_data_folder.name, reading_results_folder.name)
    logger.info("座標ファイル: %s", coord_excel_path.name)
    logger.info("スキップする問題数: %s問", skip_questions)
    
    coordinates, question_groups = parse_excel_coordinates(coord_excel_path, skip_questions)
    logger.info("座標データ: %d個のマークエリア, %d個の設問", len(coordinates), len(question_groups))
    
    # 座標データをCSVとして保存（検証用）→ 01_Results/ に配置
    try:
        csv_output_path = results_data_folder / "template_coordinates.csv"
        save_template_coordinates_debug(coordinates, csv_output_path)
    except Exception:
        pass
    
    all_questions = sorted(list(question_groups.keys()))
    question_names = {q: g['question_name'] for q, g in question_groups.items()}
    
    try:
        template_path = generate_template(coord_excel_path, results_data_folder, skip_questions)
        logger.info("テンプレート生成: %s", template_path.name)
    except Exception as e:
        logger.warning("テンプレート生成エラー: %s", e)
    
    image_files = sorted(image_folder.glob('*.jpg')) + sorted(image_folder.glob('*.png'))
    image_files = [f for f in image_files if not str(f.parent).endswith(RESULTS_FOLDER) 
                   and RESULTS_FOLDER not in f.parts]
    
    if not image_files:
        logger.error("%s に画像ファイルが見つかりません", image_folder)
        return {'success_count': 0, 'error_count': 0, 'error_files': [], 'total_count': 0, 'elapsed_time': 0}
    
    logger.info("=" * 60)
    logger.info("処理対象: %d個の画像", len(image_files))
    logger.info("=" * 60)
    
    success_count = 0
    error_count = 0
    error_files = []  # 排除候補: コーナーマーカー未検出等で処理できなかったファイル名
    all_csv_data = []
    recognition_results_list = []
    marker_cache = {}  # マーカー座標キャッシュ（Step2高速化用）
    all_kmeans_infos = []  # K-means 情報集約用
    whiteness_all = {}  # 白さキャッシュ（MarkChecker高速化用）

    # --- 並列処理 ---
    # PyInstaller frozen EXE では ProcessPoolExecutor がワーカープロセスで
    # EXE を再実行し GUI が多重起動する問題があるため、ThreadPoolExecutor を使用。
    # OpenCV/NumPy は GIL を解放するため ThreadPoolExecutor でも並列性を確保できる。
    is_frozen = getattr(sys, 'frozen', False)
    PoolExecutor = ThreadPoolExecutor if is_frozen else ProcessPoolExecutor
    max_workers = max(1, (os.cpu_count() or 1) - 1)
    total = len(image_files)
    logger.info("並列ワーカー数: %d (%s)", max_workers,
                "ThreadPool" if is_frozen else "ProcessPool")

    worker_args = [
        (str(img), str(boxed_folder), str(clean_folder), coordinates, question_groups,
         color_threshold, area_threshold, omr_mode)
        for img in image_files
    ]

    completed = 0
    with PoolExecutor(max_workers=max_workers) as executor:
        future_to_name = {
            executor.submit(_process_single_image, args): Path(args[0]).name
            for args in worker_args
        }

        for future in as_completed(future_to_name):
            # 中断チェック（新規結果の取得を停止）
            if cancel_event and cancel_event.is_set():
                # 未完了のfutureをキャンセル
                for f in future_to_name:
                    f.cancel()
                logger.info("中断されました (%d/%d件処理済み)", completed, total)
                break

            completed += 1
            fname = future_to_name[future]
            logger.info("[%d/%d] 完了: %s", completed, total, fname)

            if progress_callback:
                try:
                    progress_callback(completed, total)
                except Exception:
                    pass

            try:
                result = future.result()
                recognition_results_list.append({
                    'image': result['filename'],
                    'marks': result['marks'],
                })
                marker_cache[result['filename']] = result['marker_data']
                all_csv_data.extend(result['csv_data'])
                if result.get('kmeans_info') is not None:
                    all_kmeans_infos.append({
                        'filename': result['filename'],
                        'info': result['kmeans_info'],
                    })
                if result.get('whiteness'):
                    whiteness_all[result['filename']] = result['whiteness']
                success_count += 1
            except Exception as e:
                logger.error("処理エラー (%s): %s", fname, e)
                error_count += 1
                error_files.append(fname)

    csv_path = results_data_folder / 'coordinates.csv'
    save_coordinates_to_csv(csv_path, all_csv_data)
    logger.info("座標データCSV保存: %s", csv_path.name)

    # マーカーキャッシュをJSON保存（Step2での射影変換高速化用）
    try:
        marker_cache_path = results_data_folder / MARKER_CACHE_FILE
        with open(str(marker_cache_path), 'w', encoding='utf-8') as f:
            json.dump(marker_cache, f)
        logger.info("マーカーキャッシュ保存: %d件", len(marker_cache))
    except Exception as e:
        logger.warning("マーカーキャッシュの保存に失敗しました（Step2は再検出にフォールバックします）: %s", e)

    # 白さキャッシュをJSON保存（MarkCheckerでの白さ順ソート高速化用）
    try:
        whiteness_path = results_data_folder / WHITENESS_CACHE_FILE
        with open(str(whiteness_path), 'w', encoding='utf-8') as f:
            json.dump(whiteness_all, f, ensure_ascii=False)
        logger.info("白さキャッシュ保存: %d件", len(whiteness_all))
    except Exception as e:
        logger.warning("白さキャッシュの保存に失敗しました（MarkCheckerでフォールバック計算されます）: %s", e)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    area_str = f"{int(area_threshold * 100):03d}"
    color_str = f"{int(color_threshold * 100):03d}"
    omr_result_path = reading_results_folder / f"Mark2-Result-A{area_str}-C{color_str}-{timestamp}.xlsx"
    
    # 設問ごとの選択肢数を構築
    choice_counts = {}
    for q_no in all_questions:
        q_coords = [c for c in coordinates if c['question_no'] == q_no]
        choice_counts[q_no] = len(q_coords)
    
    save_recognition_results(omr_result_path, recognition_results_list, all_questions, question_names, choice_counts, coordinates, mark_format=mark_format)
    logger.info("OMR認識結果保存: %s", omr_result_path.name)

    # K-means HTMLレポート生成
    kmeans_report_path = None
    if omr_mode == OMR_MODE_KMEANS and all_kmeans_infos:
        try:
            kmeans_report_path = results_data_folder / f"kmeans_report_{timestamp}.html"
            generate_kmeans_report(kmeans_report_path, all_kmeans_infos)
            logger.info("K-means HTMLレポート保存: %s", kmeans_report_path.name)
        except Exception as e:
            logger.warning("K-means レポート生成エラー: %s", e)

    elapsed_time = time.time() - start_time
    
    logger.info("=" * 60)
    logger.info("処理完了: 成功 %d件 / エラー %d件", success_count, error_count)
    logger.info("実行時間: %.2f秒", elapsed_time)
    logger.info("=" * 60)
    
    return {
        'success_count': success_count,
        'error_count': error_count,
        'error_files': sorted(error_files),
        'total_count': len(image_files),
        'elapsed_time': elapsed_time,
        'kmeans_report_path': str(kmeans_report_path) if kmeans_report_path else None,
    }


# process_folderエイリアス（後方互換性のため）
process_folder = process_box_drawer
