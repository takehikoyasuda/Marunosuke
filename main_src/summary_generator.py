"""
summary_generator.py - サマリー生成モジュール

記述式採点結果をもとに、学生別得点サマリー(記述式のみモード)を
Excelファイルとして生成し、設問別の統計を出力する。
"""

from pathlib import Path
import logging
import re
import numpy as np
from PIL import Image

logger = logging.getLogger(__name__)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

from constants import (
    combine_images_to_pdf,
    RESULTS_FOLDER,
    SCORED_FOLDER,
    FINAL_REPORT_FOLDER,
    STUDENT_SUMMARY_FILE,
    EXAM_SUMMARY_FILE,
    SCORED_PDF_FILE,
    escape_excel_formula,
    get_excel_font_family,
)


def _natural_sort_key(text: str):
    """Windows Explorer 互換の自然順ソートキー

    文字列内の数値部分を int に変換し、非数値部分は小文字化して比較する。
    例: 'page2.jpg' < 'page10.jpg' (辞書順だと逆になる)
    """
    return [int(c) if c.isdigit() else c.lower()
            for c in re.split(r'(\d+)', text)]


def _build_roster_row_specs(all_files, student_id_result, roster):
    """名簿(roster)の並び順・行数に合わせて、出力行の並びを決める。

    名簿にある学籍番号の順に1行ずつ行を作る。対応する答案ファイルが
    見つかった学生は ('scored', ファイル名)、見つからなかった学生
    (未提出等)は ('absent', 学籍番号) として、名簿の行数をそのまま
    出力行数に反映する(空欄行として出力される)。名簿にない
    (または学籍番号が一致しなかった)ファイルは、末尾に自然順で追加する。

    後で名簿Excelに得点列を貼り戻すことを想定し、名簿の行順・行数を
    できるだけ壊さないようにするための処理。

    Args:
        all_files: 対象となる全ファイル名のリスト
        student_id_result: {filename: {'text': 学籍番号, ...}}
        roster: {学籍番号: 氏名} の dict（挿入順=名簿の並び順）

    Returns:
        [(kind, key), ...] のリスト。kind='scored' の場合 key はファイル名、
        kind='absent' の場合 key は学籍番号(未提出者のプレースホルダ行)。
    """
    sid_to_files = {}
    for fname, info in (student_id_result or {}).items():
        sid = (info.get('text') or '').strip()
        if sid:
            sid_to_files.setdefault(sid, []).append(fname)

    specs = []
    seen = set()
    for sid in roster.keys():
        files_for_sid = [f for f in sid_to_files.get(sid, []) if f in all_files]
        if files_for_sid:
            for fname in files_for_sid:
                specs.append(('scored', fname))
                seen.add(fname)
        else:
            specs.append(('absent', sid))

    remaining = sorted(f for f in all_files if f not in seen)
    specs.extend(('scored', f) for f in remaining)
    return specs


def _scored_pdf_order(scored_files, student_id_result, roster):
    """採点済み答案PDFにまとめる際のページ順を決める。

    名簿があれば名簿の並び順、名簿は無くても学籍番号OCRの確認結果があれば
    確認済み学籍番号の昇順で並べる。どちらも無い場合はNoneを返し、
    呼び出し側（combine_images_to_pdf）のファイル名順フォールバックに委ねる。

    Args:
        scored_files: 対象フォルダに実在するファイル名のリスト
        student_id_result: {filename: {'text': 学籍番号, ...}}
        roster: {学籍番号: 氏名} の dict（挿入順=名簿の並び順）

    Returns:
        並び替え済みのファイル名リスト。scored_filesに含まれる全件を含む
        （名簿・学籍番号に対応しないファイルは末尾に自然順で追加）。
        名簿・学籍番号OCR結果のどちらも無ければNone。
    """
    if roster:
        sid_to_files = {}
        for fname, info in (student_id_result or {}).items():
            sid = (info.get('text') or '').strip()
            if sid:
                sid_to_files.setdefault(sid, []).append(fname)
        scored_set = set(scored_files)
        ordered = []
        seen = set()
        for sid in roster.keys():
            for fname in sid_to_files.get(sid, []):
                if fname in scored_set and fname not in seen:
                    ordered.append(fname)
                    seen.add(fname)
        remaining = sorted(f for f in scored_files if f not in seen)
        ordered.extend(remaining)
        return ordered

    if student_id_result:
        def _key(fname):
            sid = (student_id_result.get(fname, {}).get('text') or '').strip()
            return (0, sid) if sid else (1, fname)
        return sorted(scored_files, key=_key)

    return None


def process_descriptive_only_summary(
    image_folder,
    descriptive_config,
    descriptive_scores,
    name_images=None,
    student_id_result=None,
    roster=None,
    output_base_folder=None,
):
    """記述のみモードのサマリー生成。

    マーク採点結果なしで、記述採点データだけから
    学生別サマリーと試験統計を生成する。

    Args:
        image_folder: 画像フォルダパス
        descriptive_config: descriptive_config dict
        descriptive_scores: {filename: {question_id: score}}
        name_images: {filename: trimmed_image_path}
        student_id_result: {filename: {'thumbnail_path','text','name'}}
            （学籍番号OCR＋人間確認済みのデータ）
        roster: {学籍番号: 氏名} の dict（挿入順=名簿の並び順）。指定時は
            出力の行順を名簿の並びに合わせる（名簿にない行は末尾に自然順で追加）。
        output_base_folder: 出力先 (None→image_folder)

    Returns:
        {'success': bool, 'stats': dict, ...} or {'success': False, 'error': str}
    """
    image_folder = Path(image_folder)
    if output_base_folder is None:
        output_base_folder = image_folder
    else:
        output_base_folder = Path(output_base_folder)

    results_folder = output_base_folder / RESULTS_FOLDER
    final_report = results_folder / FINAL_REPORT_FOLDER
    final_report.mkdir(parents=True, exist_ok=True)

    student_summary_path = final_report / STUDENT_SUMMARY_FILE
    exam_summary_path = final_report / EXAM_SUMMARY_FILE

    questions = descriptive_config.get("questions", [])
    if not questions:
        return {"success": False, "error": "記述問題が設定されていません"}

    logger.info("=" * 60)
    logger.info("サマリー生成（記述のみモード）")
    logger.info("=" * 60)
    logger.info("✓ 記述問題: %d問", len(questions))
    logger.info("✓ 対象画像: %d件", len(descriptive_scores))
    if name_images:
        logger.info("✓ 氏名欄画像: %d枚", len(name_images))
    has_student_id_result = student_id_result is not None and len(student_id_result) > 0
    has_roster_names = has_student_id_result and any(v.get('name') for v in student_id_result.values())
    if has_student_id_result:
        logger.info("✓ 学籍番号OCR確認済み: %d枚", len(student_id_result))
    logger.info("")

    try:
        # --- 学生別サマリー ---
        wb_student = Workbook()
        ws = wb_student.active
        ws.title = "学生別サマリー"

        # ヘッダースタイル
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # ヘッダー構築
        headers = ["No.", "ファイル名"]
        if name_images:
            headers.append("氏名欄")
        if has_student_id_result:
            headers.append("学籍番号欄")
            headers.append("学籍番号(確認済み)")
            if has_roster_names:
                headers.append("氏名候補(名簿照合)")
        for q in questions:
            headers.append(f"{q['name']} ({q['max_score']})")
        headers.append("合計")
        full_score = sum(q["max_score"] for q in questions)
        headers.append(f"配点計 ({full_score})")

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        # データ行（名簿がある場合は名簿の並び順・行数、無ければファイル名の自然順）
        if roster:
            row_specs = _build_roster_row_specs(list(descriptive_scores.keys()), student_id_result, roster)
        else:
            row_specs = [('scored', f) for f in sorted(descriptive_scores.keys())]
        # 設問別統計は実際に採点された(未提出でない)ファイルのみを対象にする
        scored_filenames = [key for kind, key in row_specs if kind == 'scored']
        totals = []
        for row_idx, (kind, key) in enumerate(row_specs, 2):
            is_absent = kind == 'absent'
            fname = None if is_absent else key
            scores_for_file = {} if is_absent else descriptive_scores.get(fname, {})
            col = 1
            ws.cell(row=row_idx, column=col, value=row_idx - 1).border = thin_border
            col += 1
            fname_display = "(未提出)" if is_absent else escape_excel_formula(fname)
            ws.cell(row=row_idx, column=col, value=fname_display).border = thin_border
            col += 1

            if name_images:
                name_path = None if is_absent else name_images.get(fname)
                if name_path and Path(name_path).exists():
                    try:
                        img = XlImage(str(name_path))
                        img.width = 120
                        img.height = 30
                        ws.add_image(img, get_column_letter(col) + str(row_idx))
                    except Exception:
                        pass
                ws.cell(row=row_idx, column=col).border = thin_border
                col += 1

            if has_student_id_result:
                sid_info = {} if is_absent else student_id_result.get(fname, {})
                sid_thumb = sid_info.get('thumbnail_path')
                if sid_thumb and Path(sid_thumb).exists():
                    try:
                        img = XlImage(str(sid_thumb))
                        img.width = 120
                        img.height = 30
                        ws.add_image(img, get_column_letter(col) + str(row_idx))
                    except Exception:
                        pass
                ws.cell(row=row_idx, column=col).border = thin_border
                col += 1

                # 未提出者は名簿から学籍番号がすでに分かっているのでそのまま表示する
                sid_text = key if is_absent else (sid_info.get('text') or '')
                ws.cell(row=row_idx, column=col, value=escape_excel_formula(sid_text)).border = thin_border
                ws.cell(row=row_idx, column=col).alignment = center
                col += 1

                if has_roster_names:
                    name_text = roster.get(key, '') if is_absent else (sid_info.get('name') or '')
                    ws.cell(row=row_idx, column=col, value=escape_excel_formula(name_text)).border = thin_border
                    ws.cell(row=row_idx, column=col).alignment = center
                    col += 1

            student_total = None if is_absent else 0
            for q in questions:
                if is_absent:
                    ws.cell(row=row_idx, column=col).border = thin_border
                else:
                    sc = scores_for_file.get(q["id"], 0)
                    if sc is None:
                        sc = 0
                    ws.cell(row=row_idx, column=col, value=sc).border = thin_border
                    ws.cell(row=row_idx, column=col).alignment = center
                    student_total += sc
                col += 1

            if not is_absent:
                ws.cell(row=row_idx, column=col, value=student_total).border = thin_border
                ws.cell(row=row_idx, column=col).alignment = center
                ws.cell(row=row_idx, column=col).font = Font(bold=True)
                totals.append(student_total)
            else:
                ws.cell(row=row_idx, column=col).border = thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=full_score).border = thin_border
            ws.cell(row=row_idx, column=col).alignment = center

        # 列幅調整
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 30
        img_col_idx = 3
        if name_images:
            ws.column_dimensions[get_column_letter(img_col_idx)].width = 18
            img_col_idx += 1
        if has_student_id_result:
            ws.column_dimensions[get_column_letter(img_col_idx)].width = 18
            img_col_idx += 1
            ws.column_dimensions[get_column_letter(img_col_idx)].width = 14
            img_col_idx += 1
            if has_roster_names:
                ws.column_dimensions[get_column_letter(img_col_idx)].width = 16
                img_col_idx += 1
        if name_images or has_student_id_result:
            ws.row_dimensions[1].height = 20
            for r in range(2, len(row_specs) + 2):
                ws.row_dimensions[r].height = 25

        wb_student.save(str(student_summary_path))
        logger.info("✓ 学生別サマリー: %s", student_summary_path.name)

        # --- 試験統計 ---
        totals_arr = np.array(totals) if totals else np.array([0])
        exam_stats = {
            "受験者数": len(totals),
            "満点": full_score,
            "平均点": float(np.mean(totals_arr)) if totals else 0.0,
            "標準偏差": float(np.std(totals_arr, ddof=1)) if len(totals) > 1 else 0.0,
            "最高点": int(np.max(totals_arr)) if totals else 0,
            "最低点": int(np.min(totals_arr)) if totals else 0,
        }

        wb_exam = Workbook()
        ws_exam = wb_exam.active
        ws_exam.title = "試験統計"
        stat_items = [
            ("受験者数", exam_stats["受験者数"]),
            ("満点", exam_stats["満点"]),
            ("平均点", f"{exam_stats['平均点']:.2f}"),
            ("標準偏差", f"{exam_stats['標準偏差']:.2f}"),
            ("最高点", exam_stats["最高点"]),
            ("最低点", exam_stats["最低点"]),
        ]
        ws_exam.cell(row=1, column=1, value="項目").font = header_font
        ws_exam.cell(row=1, column=2, value="値").font = header_font
        for r, (k, v) in enumerate(stat_items, 2):
            ws_exam.cell(row=r, column=1, value=k)
            ws_exam.cell(row=r, column=2, value=v)
        ws_exam.column_dimensions["A"].width = 20
        ws_exam.column_dimensions["B"].width = 20

        # 設問別統計シート
        ws_q = wb_exam.create_sheet("設問別統計")
        q_headers = ["設問", "配点", "平均", "標準偏差", "最高", "最低", "正答率(%)"]
        for ci, h in enumerate(q_headers, 1):
            cell = ws_q.cell(row=1, column=ci, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center

        for qi, q in enumerate(questions, 2):
            qid = q["id"]
            q_scores = [
                descriptive_scores.get(f, {}).get(qid, 0) or 0
                for f in scored_filenames
            ]
            q_arr = np.array(q_scores) if q_scores else np.array([0])
            ws_q.cell(row=qi, column=1, value=q["name"])
            ws_q.cell(row=qi, column=2, value=q["max_score"]).alignment = center
            ws_q.cell(row=qi, column=3, value=f"{np.mean(q_arr):.2f}").alignment = center
            sd = float(np.std(q_arr, ddof=1)) if len(q_scores) > 1 else 0.0
            ws_q.cell(row=qi, column=4, value=f"{sd:.2f}").alignment = center
            ws_q.cell(row=qi, column=5, value=int(np.max(q_arr))).alignment = center
            ws_q.cell(row=qi, column=6, value=int(np.min(q_arr))).alignment = center
            rate = float(np.mean(q_arr)) / q["max_score"] * 100 if q["max_score"] > 0 else 0.0
            ws_q.cell(row=qi, column=7, value=f"{rate:.1f}").alignment = center

        wb_exam.save(str(exam_summary_path))
        logger.info("✓ 試験統計: %s", exam_summary_path.name)

        # 統合PDF（名簿があれば名簿順、無くても学籍番号OCR結果があれば
        # 確認済み学籍番号の昇順、どちらも無ければファイル名順で並べる）
        scored_folder = results_folder / SCORED_FOLDER
        scored_pdf_path = final_report / SCORED_PDF_FILE
        scored_pdf_result_path = None
        scored_pdf_error = None
        if scored_folder.exists():
            try:
                scored_files = sorted(
                    f.name for f in scored_folder.iterdir()
                    if f.suffix.lower() in ('.jpg', '.png')
                )
                ordered_filenames = _scored_pdf_order(scored_files, student_id_result, roster)
                generated = combine_images_to_pdf(scored_folder, scored_pdf_path, ordered_filenames=ordered_filenames)
                if generated:
                    scored_pdf_result_path = str(scored_pdf_path)
                    logger.info("✓ 統合PDF: %s", scored_pdf_path.name)
                else:
                    logger.info("統合PDF: 採点済み答案の画像が無いためスキップしました")
            except Exception as pdf_e:
                scored_pdf_error = str(pdf_e)
                logger.warning("統合PDF生成エラー: %s", pdf_e)

        logger.info("")
        logger.info("=" * 60)
        logger.info("サマリー生成完了（記述のみモード）")
        logger.info("=" * 60)
        logger.info("✓ 学生別サマリー: %s", student_summary_path.name)
        logger.info("✓ 試験統計: %s", exam_summary_path.name)

        result = {
            "success": True,
            "student_summary_path": str(student_summary_path),
            "exam_summary_path": str(exam_summary_path),
            "scored_pdf_path": scored_pdf_result_path,
            "scored_pdf_error": scored_pdf_error,
            "stats": exam_stats,
        }
        return result

    except Exception as e:
        logger.error("エラー: %s", e, exc_info=True)
        return {"success": False, "error": str(e)}
