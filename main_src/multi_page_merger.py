#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
multi_page_merger.py — 複数ページ答案の取込管理・監査・集計統合。

アプリへ追加したPDF・画像を取込バッチとして管理し、展開された各画像を
「学籍番号 × 試験ページ」で関連付ける。試験ページとPDF内ページを区別し、
同じ試験ページが複数PDFへ分割された運用にも対応する。

1人の学生が複数ページ（例: 3枚、各ページに異なる設問）を提出する運用を前提に、
「同じページ番号の答案だけをまとめたバッチ」ごとに独立して生成した集計Excel
（記述のみモード, summary_generator.process_descriptive_only_summary の出力）
を、学籍番号OCRで確認済みの学籍番号をキーに1つに統合する。

既存の集計Excelを後から読み込んで突合する補助ツールも、後方互換のため維持する。
"""

import json
import logging
import re
import shutil
import statistics
import tkinter as tk
import uuid
from dataclasses import asdict, dataclass, field
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog, ttk
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from constants import get_ui_font_family, get_ui_font_size, fit_window_to_content

logger = logging.getLogger(__name__)

UI_FONT = get_ui_font_family()

# process_descriptive_only_summary が出力する列見出しのうち、設問列ではないもの
NON_QUESTION_COLUMNS = {
    'No.', 'ファイル名', '氏名欄', '学籍番号欄', '学籍番号(確認済み)', '氏名候補(名簿照合)', '合計',
}
STUDENT_ID_COLUMN = '学籍番号(確認済み)'
TOTAL_COLUMN = '合計'
MAX_SCORE_PREFIX = '配点計 ('
MULTI_PAGE_MANIFEST_FILE = 'multi_page_manifest.json'
MULTI_PAGE_PROJECT_POINTER_FILE = 'multi_page_project.json'
SHARED_LAYOUT_FOLDER = 'multi_page_shared_layout'
SHARED_LAYOUT_FILES = (
    'descriptive_config.json',
    'name_area_config.json',
    'student_id_area_config.json',
)
COMBINED_STUDENT_SUMMARY_FILE = '001_all_pages_student_summary.xlsx'


@dataclass
class AnswerPage:
    """学生1人分の答案1ページ。

    ``exam_page`` は試験用紙上のページ番号、``source_page`` は取込元PDF内の
    ページ番号であり、両者を明確に区別する。
    """

    image_id: str
    image_path: str
    exam_page: int
    batch_id: str
    source_path: str
    source_page: int
    student_id: str = ''
    status: str = 'imported'


@dataclass
class ImportBatch:
    """同じ試験ページとして一度に追加したPDF・画像のまとまり。"""

    batch_id: str
    exam_page: int
    source_paths: List[str] = field(default_factory=list)
    answer_pages: List[AnswerPage] = field(default_factory=list)


@dataclass
class MultiPageAudit:
    """複数ページ答案の関連付け監査結果。"""

    exam_pages: List[int] = field(default_factory=list)
    associations: Dict[str, Dict[str, str]] = field(default_factory=dict)
    missing_pages: Dict[str, List[int]] = field(default_factory=dict)
    duplicates: Dict[str, Dict[str, List[str]]] = field(default_factory=dict)
    unmatched_image_ids: List[str] = field(default_factory=list)
    unexpected_student_ids: List[str] = field(default_factory=list)
    unexpected_exam_pages: Dict[str, int] = field(default_factory=dict)
    shared_layout: bool = False
    active_exam_page: Optional[int] = None

    @property
    def is_ready(self) -> bool:
        """不足・重複・未関連付けがなく、採点へ進める状態か返す。"""
        return not any((
            self.missing_pages,
            self.duplicates,
            self.unmatched_image_ids,
            self.unexpected_student_ids,
            self.unexpected_exam_pages,
        ))


def create_import_batch(
    exam_page: int,
    sources: Dict[str, List[str]],
    batch_id: Optional[str] = None,
) -> ImportBatch:
    """展開済み画像を同じ試験ページの取込バッチとして登録する。

    Args:
        exam_page: 試験用紙上のページ番号（1始まり）。
        sources: ``{元PDFまたは画像のパス: [展開済み画像パス, ...]}``。
            同じ試験ページの複数PDFを一度のバッチへ含められる。
        batch_id: 復元・テスト時に固定IDを使う場合のみ指定する。
    """
    if not isinstance(exam_page, int) or isinstance(exam_page, bool) or exam_page < 1:
        raise ValueError("試験ページ番号は1以上の整数で指定してください。")
    resolved_batch_id = batch_id or str(uuid.uuid4())
    answer_pages = []
    for source_path, image_paths in sources.items():
        for source_page, image_path in enumerate(image_paths, 1):
            answer_pages.append(AnswerPage(
                image_id=str(uuid.uuid4()),
                image_path=str(image_path),
                exam_page=exam_page,
                batch_id=resolved_batch_id,
                source_path=str(source_path),
                source_page=source_page,
            ))
    return ImportBatch(
        batch_id=resolved_batch_id,
        exam_page=exam_page,
        source_paths=[str(path) for path in sources],
        answer_pages=answer_pages,
    )


def import_files_as_batch(
    exam_page: int,
    source_paths: List[str],
    managed_root: str,
    batch_id: Optional[str] = None,
) -> ImportBatch:
    """PDF・画像をアプリ管理領域へ取り込み、取込バッチを作る。

    元ファイルは変更しない。PDFは学生1人につき1 PDFページとしてPNGへ展開し、
    画像はそのまま管理領域へコピーする。途中で失敗した場合は、この呼び出しで
    新規作成したバッチディレクトリだけを除去する。
    """
    from constants import extract_pdf_to_images

    if not isinstance(exam_page, int) or isinstance(exam_page, bool) or exam_page < 1:
        raise ValueError("試験ページ番号は1以上の整数で指定してください。")
    if not source_paths:
        raise ValueError("取り込むPDFまたは画像を指定してください。")
    resolved_sources = [str(Path(path).resolve()) for path in source_paths]
    if len(resolved_sources) != len(set(resolved_sources)):
        raise ValueError("同じファイルが複数回選択されています。")

    resolved_batch_id = batch_id or str(uuid.uuid4())
    root = Path(managed_root)
    batch_dir = root / resolved_batch_id
    if batch_dir.exists():
        raise FileExistsError(f"取込バッチが既に存在します: {resolved_batch_id}")
    batch_dir.mkdir(parents=True)

    answer_pages: List[AnswerPage] = []
    try:
        for source_index, source_text in enumerate(resolved_sources, 1):
            source = Path(source_text)
            if not source.is_file():
                raise FileNotFoundError(f"取込ファイルが見つかりません: {source}")
            source_dir = batch_dir / f"source_{source_index:03d}"
            source_dir.mkdir()
            if source.suffix.lower() == '.pdf':
                extract_pdf_to_images(str(source), output_folder=source_dir)
                image_paths = sorted(source_dir.glob('*.png'))
            elif source.suffix.lower() in {'.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff'}:
                destination = source_dir / source.name
                shutil.copy2(source, destination)
                image_paths = [destination]
            else:
                raise ValueError(f"対応していないファイル形式です: {source.name}")

            for source_page, image_path in enumerate(image_paths, 1):
                answer_pages.append(AnswerPage(
                    image_id=str(uuid.uuid4()),
                    image_path=str(image_path),
                    exam_page=exam_page,
                    batch_id=resolved_batch_id,
                    source_path=str(source),
                    source_page=source_page,
                ))
    except Exception:
        shutil.rmtree(batch_dir, ignore_errors=True)
        raise

    return ImportBatch(
        batch_id=resolved_batch_id,
        exam_page=exam_page,
        source_paths=resolved_sources,
        answer_pages=answer_pages,
    )


def prepare_exam_page_workspace(
    answer_pages: List[AnswerPage],
    exam_page: int,
    workspace_root: str,
    project_folder: str,
    shared_layout: bool = False,
) -> str:
    """選択した試験ページの答案画像を既存採点フロー用フォルダへ同期する。"""
    selected = [
        answer for answer in answer_pages
        if answer.exam_page == exam_page and getattr(answer, 'status', 'imported') != 'excluded'
    ]
    if not selected:
        raise ValueError(f"試験ページ {exam_page} に答案画像がありません。")

    workspace = Path(workspace_root) / f"page_{exam_page:03d}"
    workspace.mkdir(parents=True, exist_ok=True)
    expected_names = set()
    for answer in selected:
        source = Path(answer.image_path)
        if not source.is_file():
            raise FileNotFoundError(f"展開済み答案画像が見つかりません: {source}")
        suffix = source.suffix.lower() or '.png'
        filename = f"{answer.image_id}{suffix}"
        expected_names.add(filename)
        destination = workspace / filename
        if (not destination.exists()
                or destination.stat().st_mtime < source.stat().st_mtime
                or destination.stat().st_size != source.stat().st_size):
            shutil.copy2(source, destination)

    supported = {'.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff'}
    for existing in workspace.iterdir():
        if (existing.is_file() and existing.suffix.lower() in supported
                and existing.name not in expected_names):
            existing.unlink()

    pointer = {
        'version': 1,
        'project_folder': str(Path(project_folder).resolve()),
        'exam_page': exam_page,
        'shared_layout': shared_layout,
    }
    with (workspace / MULTI_PAGE_PROJECT_POINTER_FILE).open('w', encoding='utf-8') as handle:
        json.dump(pointer, handle, ensure_ascii=False, indent=2)
    if shared_layout:
        copied = apply_shared_layout_settings(project_folder, str(workspace))
        if copied:
            pointer['shared_settings_applied'] = True
            with (workspace / MULTI_PAGE_PROJECT_POINTER_FILE).open('w', encoding='utf-8') as handle:
                json.dump(pointer, handle, ensure_ascii=False, indent=2)
    return str(workspace)


def resolve_multi_page_project_folder(folder: str) -> str:
    """ページ作業フォルダなら元の複数ページプロジェクトを返す。"""
    pointer_path = Path(folder) / MULTI_PAGE_PROJECT_POINTER_FILE
    if not pointer_path.exists():
        return str(Path(folder))
    try:
        with pointer_path.open('r', encoding='utf-8') as handle:
            payload = json.load(handle)
        project_folder = Path(payload['project_folder'])
        if payload.get('version') == 1 and project_folder.is_dir():
            return str(project_folder)
    except (OSError, ValueError, KeyError, TypeError, json.JSONDecodeError):
        pass
    return str(Path(folder))


def get_exam_page_for_workspace(folder: str) -> Optional[int]:
    """試験ページの作業フォルダなら、そのページ番号を返す。単一ページ案件はNone。"""
    pointer_path = Path(folder) / MULTI_PAGE_PROJECT_POINTER_FILE
    if not pointer_path.exists():
        return None
    try:
        with pointer_path.open('r', encoding='utf-8') as handle:
            payload = json.load(handle)
        return int(payload['exam_page'])
    except (OSError, ValueError, KeyError, TypeError, json.JSONDecodeError):
        return None


def resolve_roster_config_path(image_folder: str) -> str:
    """名簿設定ファイルのパスを返す。

    名簿はページに依存しない（試験全体で共通）ため、複数ページ案件では
    案件フォルダ直下の1箇所に統一する。単一ページ案件では、指定フォルダ
    直下をそのまま使う（従来通り）。

    複数ページ案件で、まだ案件レベルの名簿が無い場合は、いずれかの
    ページに個別保存されていた名簿があれば引き上げて救済する
    （旧バージョンでページごとに保存されていた案件との互換用）。
    """
    from constants import RESULTS_DATA_FOLDER, RESULTS_FOLDER
    from roster_config import ROSTER_CONFIG_FILE

    project_folder = Path(resolve_multi_page_project_folder(image_folder))
    project_path = project_folder / RESULTS_FOLDER / RESULTS_DATA_FOLDER / ROSTER_CONFIG_FILE
    if project_folder == Path(image_folder):
        return str(project_path)  # 単一ページ案件

    if not project_path.is_file():
        workspace_root = project_folder / '_multi_page_pages'
        if workspace_root.exists():
            for page_workspace in sorted(workspace_root.glob('page_*')):
                candidate = page_workspace / RESULTS_FOLDER / RESULTS_DATA_FOLDER / ROSTER_CONFIG_FILE
                if candidate.is_file():
                    project_path.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(candidate, project_path)
                    break
    return str(project_path)


def shared_layout_was_applied(folder: str) -> bool:
    """ページ準備時にプロジェクト共通設定が適用済みか返す。"""
    pointer_path = Path(folder) / MULTI_PAGE_PROJECT_POINTER_FILE
    try:
        with pointer_path.open('r', encoding='utf-8') as handle:
            return bool(json.load(handle).get('shared_settings_applied'))
    except (OSError, ValueError, TypeError, json.JSONDecodeError):
        return False


def _shared_layout_dir(project_folder: str) -> Path:
    from constants import RESULTS_DATA_FOLDER, RESULTS_FOLDER
    return Path(project_folder) / RESULTS_FOLDER / RESULTS_DATA_FOLDER / SHARED_LAYOUT_FOLDER


def publish_shared_layout_settings(page_workspace: str) -> List[str]:
    """ページで完了した初期設定を複数ページ共通設定として保存する。"""
    from constants import RESULTS_DATA_FOLDER, RESULTS_FOLDER

    pointer_path = Path(page_workspace) / MULTI_PAGE_PROJECT_POINTER_FILE
    if not pointer_path.exists():
        return []
    with pointer_path.open('r', encoding='utf-8') as handle:
        pointer = json.load(handle)
    if not pointer.get('shared_layout'):
        return []

    source_dir = Path(page_workspace) / RESULTS_FOLDER / RESULTS_DATA_FOLDER
    destination_dir = _shared_layout_dir(pointer['project_folder'])
    destination_dir.mkdir(parents=True, exist_ok=True)
    copied = []
    for filename in SHARED_LAYOUT_FILES:
        source = source_dir / filename
        if source.is_file():
            shutil.copy2(source, destination_dir / filename)
            copied.append(filename)
    return copied


def bootstrap_shared_layout_settings(project_folder: str, workspace_root: str) -> List[str]:
    """既存ページ設定から共通設定を自動作成する。

    初回ページの設定完了時に共通設定の公開が行われなかった旧経路も救済する。
    既に共通設定があれば何もしない。
    """
    from constants import RESULTS_DATA_FOLDER, RESULTS_FOLDER

    shared_dir = _shared_layout_dir(project_folder)
    if (shared_dir / 'descriptive_config.json').is_file():
        return []
    root = Path(workspace_root)
    if not root.exists():
        return []
    for page_workspace in sorted(root.glob('page_*')):
        data_dir = page_workspace / RESULTS_FOLDER / RESULTS_DATA_FOLDER
        if not (data_dir / 'descriptive_config.json').is_file():
            continue
        shared_dir.mkdir(parents=True, exist_ok=True)
        copied = []
        for filename in SHARED_LAYOUT_FILES:
            source = data_dir / filename
            if source.is_file():
                shutil.copy2(source, shared_dir / filename)
                copied.append(filename)
        return copied
    return []


def apply_shared_layout_settings(project_folder: str, page_workspace: str) -> List[str]:
    """保存済みの共通初期設定をページ作業フォルダへ適用する。"""
    from constants import RESULTS_DATA_FOLDER, RESULTS_FOLDER

    source_dir = _shared_layout_dir(project_folder)
    if not source_dir.exists():
        return []
    destination_dir = Path(page_workspace) / RESULTS_FOLDER / RESULTS_DATA_FOLDER
    destination_dir.mkdir(parents=True, exist_ok=True)
    copied = []
    for filename in SHARED_LAYOUT_FILES:
        source = source_dir / filename
        if source.is_file():
            shutil.copy2(source, destination_dir / filename)
            copied.append(filename)
    return copied


def audit_answer_pages(
    answer_pages: List[AnswerPage],
    expected_pages: Optional[List[int]] = None,
    roster: Optional[Dict[str, str]] = None,
) -> MultiPageAudit:
    """答案を「学籍番号 × 試験ページ」で関連付け、問題を検出する。

    名簿が指定された場合は名簿の全員を監査対象とし、名簿外の学籍番号も検出する。
    名簿がない場合は、いずれかの答案で確認できた学籍番号を対象とする。
    同一ページ内で同じ学籍番号が複数ある場合は、誤った答案を採用しないよう、その
    ページの関連付けを確定しない。
    """
    answer_pages = [a for a in answer_pages if getattr(a, 'status', 'imported') != 'excluded']

    if expected_pages is None:
        pages = sorted({answer.exam_page for answer in answer_pages})
    else:
        pages = list(expected_pages)
        if (any(not isinstance(page, int) or isinstance(page, bool) or page < 1 for page in pages)
                or len(pages) != len(set(pages))):
            raise ValueError("試験ページ番号は重複しない1以上の整数で指定してください。")

    per_page: Dict[int, Dict[str, List[AnswerPage]]] = {page: {} for page in pages}
    unmatched = []
    duplicates: Dict[str, Dict[str, List[str]]] = {}
    observed_ids = set()
    unexpected_exam_pages = {}

    for answer in answer_pages:
        student_id = str(answer.student_id or '').strip()
        if not student_id:
            unmatched.append(answer.image_id)
            continue
        observed_ids.add(student_id)
        if answer.exam_page not in per_page:
            unexpected_exam_pages[answer.image_id] = answer.exam_page
            continue
        per_page[answer.exam_page].setdefault(student_id, []).append(answer)

    for page, by_student in per_page.items():
        page_duplicates = {
            sid: [answer.image_id for answer in answers]
            for sid, answers in by_student.items() if len(answers) > 1
        }
        if page_duplicates:
            duplicates[str(page)] = page_duplicates

    expected_ids = list(roster.keys()) if roster is not None else sorted(observed_ids)
    expected_set = set(expected_ids)
    unexpected = sorted(observed_ids - expected_set) if roster is not None else []
    associations: Dict[str, Dict[str, str]] = {}
    missing_pages: Dict[str, List[str]] = {}

    for student_id in expected_ids:
        student_pages: Dict[str, str] = {}
        missing = []
        for page in pages:
            answers = per_page[page].get(student_id, [])
            if len(answers) == 1:
                student_pages[str(page)] = answers[0].image_id
            else:
                missing.append(page)
        associations[student_id] = student_pages
        if missing:
            missing_pages[student_id] = missing

    return MultiPageAudit(
        exam_pages=pages,
        associations=associations,
        missing_pages=missing_pages,
        duplicates=duplicates,
        unmatched_image_ids=unmatched,
        unexpected_student_ids=unexpected,
        unexpected_exam_pages=unexpected_exam_pages,
    )


def toggle_answer_status(batches: List[ImportBatch], image_id: str) -> Optional[AnswerPage]:
    """指定したAnswerPageの除外状態を切り替える（保存は呼び出し側の責務）。"""
    for batch in batches:
        for answer in batch.answer_pages:
            if answer.image_id == image_id:
                answer.status = 'imported' if getattr(answer, 'status', 'imported') == 'excluded' else 'excluded'
                return answer
    return None


def save_multi_page_manifest(
    batches: List[ImportBatch],
    audit: MultiPageAudit,
    output_path: str,
) -> None:
    """取込バッチと監査結果をセッション復元用JSONとして保存する。"""
    payload = {
        'version': 1,
        'batches': [asdict(batch) for batch in batches],
        'audit': {
            **asdict(audit),
            'is_ready': audit.is_ready,
        },
    }
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8') as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def load_multi_page_manifest(input_path: str) -> Tuple[List[ImportBatch], MultiPageAudit]:
    """保存済みの取込バッチと監査結果を読み込む。"""
    with Path(input_path).open('r', encoding='utf-8') as handle:
        payload = json.load(handle)
    if payload.get('version') != 1:
        raise ValueError("対応していない複数ページ管理データです。")

    batches = []
    for item in payload.get('batches', []):
        answers = [AnswerPage(**answer) for answer in item.get('answer_pages', [])]
        batches.append(ImportBatch(
            batch_id=item['batch_id'],
            exam_page=item['exam_page'],
            source_paths=item.get('source_paths', []),
            answer_pages=answers,
        ))
    audit_data = dict(payload.get('audit', {}))
    audit_data.pop('is_ready', None)
    return batches, MultiPageAudit(**audit_data)


def get_multi_page_project_status(folder: str) -> Optional[dict]:
    """トップ画面表示用に複数ページ案件の進捗を集計する。"""
    from constants import FINAL_REPORT_FOLDER, RESULTS_DATA_FOLDER, RESULTS_FOLDER, STUDENT_SUMMARY_FILE

    project_folder = Path(resolve_multi_page_project_folder(folder))
    manifest = project_folder / RESULTS_FOLDER / RESULTS_DATA_FOLDER / MULTI_PAGE_MANIFEST_FILE
    if not manifest.is_file():
        return None
    batches, audit = load_multi_page_manifest(str(manifest))
    answers_by_page = {
        page: [answer for batch in batches for answer in batch.answer_pages
               if answer.exam_page == page and getattr(answer, 'status', 'imported') != 'excluded']
        for page in audit.exam_pages
    }
    page_states = []
    for page in audit.exam_pages:
        workspace = project_folder / '_multi_page_pages' / f'page_{page:03d}'
        data_dir = workspace / RESULTS_FOLDER / RESULTS_DATA_FOLDER
        config_path = data_dir / 'descriptive_config.json'
        scores_path = data_dir / 'descriptive_scores.json'
        summary_path = workspace / RESULTS_FOLDER / FINAL_REPORT_FOLDER / STUDENT_SUMMARY_FILE
        question_ids = []
        if config_path.is_file():
            try:
                config = json.loads(config_path.read_text(encoding='utf-8'))
                question_ids = [q['id'] for q in config.get('questions', [])]
            except (OSError, ValueError, KeyError, TypeError, json.JSONDecodeError):
                question_ids = []
        scores = {}
        if scores_path.is_file():
            try:
                scores = json.loads(scores_path.read_text(encoding='utf-8')).get('scores', {})
            except (OSError, ValueError, TypeError, json.JSONDecodeError):
                scores = {}
        expected_files = {f"{answer.image_id}{Path(answer.image_path).suffix.lower() or '.png'}"
                          for answer in answers_by_page[page]}
        completed = sum(
            1 for filename in expected_files
            if question_ids and all(qid in scores.get(filename, {}) for qid in question_ids)
        )
        answer_count = len(answers_by_page[page])
        if summary_path.is_file():
            state = '集計済み'
        elif answer_count and completed == answer_count:
            state = '採点完了'
        elif completed:
            state = f'採点中 {completed}/{answer_count}'
        elif question_ids:
            state = '初期設定済み'
        elif answer_count:
            state = '取込済み'
        else:
            state = '未取込'
        page_states.append({
            'page': page,
            'answers': answer_count,
            'completed_answers': completed,
            'state': state,
            'workspace': str(workspace),
            'summary_path': str(summary_path),
        })
    return {
        'project_folder': str(project_folder),
        'active_page': audit.active_exam_page,
        'shared_layout': audit.shared_layout,
        'pages': page_states,
        'combined_summary_path': str(
            project_folder / RESULTS_FOLDER / FINAL_REPORT_FOLDER
            / COMBINED_STUDENT_SUMMARY_FILE
        ),
    }


def generate_combined_multi_page_summary(project_folder: str) -> dict:
    """全ページの学生別サマリーを学籍番号で統合する。"""
    from constants import FINAL_REPORT_FOLDER, RESULTS_FOLDER

    status = get_multi_page_project_status(project_folder)
    if not status:
        return {'success': False, 'error': '複数ページ案件が見つかりません。'}
    missing = [item['page'] for item in status['pages'] if not Path(item['summary_path']).is_file()]
    if missing:
        return {'success': False, 'pending_pages': missing}
    pages = []
    missing_student_id_pages = []
    try:
        for item in status['pages']:
            try:
                pages.append(read_page_summary(item['summary_path'], f"ページ{item['page']}"))
            except ValueError:
                # 「学籍番号(確認済み)」列が無い＝そのページで学籍番号OCRの
                # 確認をしていない集計Excel。全ページ分まとめて原因を示す
                # ため、最初の1件で止めずに全ページを走査する。
                missing_student_id_pages.append(item['page'])
        if missing_student_id_pages:
            return {'success': False, 'missing_student_id_pages': missing_student_id_pages}
        merged, warnings = merge_page_summaries(pages)
        output = (
            Path(status['project_folder']) / RESULTS_FOLDER / FINAL_REPORT_FOLDER
            / COMBINED_STUDENT_SUMMARY_FILE
        )
        write_merged_excel(merged, pages, str(output))
        return {
            'success': True,
            'output_path': str(output),
            'student_count': len(merged),
            'warnings': warnings,
        }
    except Exception as exc:
        return {'success': False, 'error': str(exc)}


def activate_exam_page(project_folder: str, exam_page: int) -> str:
    """指定ページを採点対象として準備し、現在ページを保存する。"""
    from constants import RESULTS_DATA_FOLDER, RESULTS_FOLDER

    project = Path(resolve_multi_page_project_folder(project_folder))
    manifest = project / RESULTS_FOLDER / RESULTS_DATA_FOLDER / MULTI_PAGE_MANIFEST_FILE
    batches, audit = load_multi_page_manifest(str(manifest))
    if exam_page not in audit.exam_pages:
        raise ValueError(f"試験ページ {exam_page} は登録されていません。")
    answers = [answer for batch in batches for answer in batch.answer_pages]
    workspace_root = project / '_multi_page_pages'
    if audit.shared_layout:
        bootstrap_shared_layout_settings(str(project), str(workspace_root))
    workspace = prepare_exam_page_workspace(
        answers, exam_page, str(workspace_root), str(project),
        shared_layout=audit.shared_layout,
    )
    audit.active_exam_page = exam_page
    save_multi_page_manifest(batches, audit, str(manifest))
    return workspace


@dataclass
class PageSummary:
    """1ページ分の集計Excelから読み取った内容。"""
    label: str
    source_path: str
    question_columns: List[str] = field(default_factory=list)
    rows_by_student_id: Dict[str, Dict] = field(default_factory=dict)
    max_score: Optional[float] = None
    unmatched_count: int = 0


def read_page_summary(excel_path: str, label: str) -> PageSummary:
    """記述のみモードの集計Excelを読み込み、PageSummaryを構築する。

    Raises:
        ValueError: 「学籍番号(確認済み)」列が見つからない場合
            （学籍番号OCRで確認されていないExcelは統合できない）。
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    col_index: Dict[str, int] = {}
    for idx, name in enumerate(header):
        if name is not None:
            col_index[name] = idx

    if STUDENT_ID_COLUMN not in col_index:
        raise ValueError(
            f"「{STUDENT_ID_COLUMN}」列が見つかりません: {excel_path}\n"
            "学籍番号OCRで確認済みの集計Excelのみ統合できます。"
        )
    sid_idx = col_index[STUDENT_ID_COLUMN]
    total_idx = col_index.get(TOTAL_COLUMN)

    question_columns = [
        name for name in header
        if name is not None and name not in NON_QUESTION_COLUMNS and not name.startswith(MAX_SCORE_PREFIX)
    ]
    question_indices = {name: header.index(name) for name in question_columns}

    max_score = None
    for name in header:
        if isinstance(name, str) and name.startswith(MAX_SCORE_PREFIX):
            m = re.search(r'\(([\d.]+)\)', name)
            if m:
                max_score = float(m.group(1))
            break

    rows_by_student_id: Dict[str, Dict] = {}
    unmatched_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[sid_idx] is None or not str(row[sid_idx]).strip():
            unmatched_count += 1
            continue
        student_id = str(row[sid_idx]).strip()
        entry = {name: row[idx] for name, idx in question_indices.items()}
        entry[TOTAL_COLUMN] = row[total_idx] if total_idx is not None else None
        rows_by_student_id[student_id] = entry

    return PageSummary(
        label=label,
        source_path=excel_path,
        question_columns=question_columns,
        rows_by_student_id=rows_by_student_id,
        max_score=max_score,
        unmatched_count=unmatched_count,
    )


def merge_page_summaries(pages: List[PageSummary]) -> Tuple[Dict[str, Dict], List[str]]:
    """複数ページ分のPageSummaryを、学籍番号をキーに1つに統合する。

    Returns:
        (統合結果 {学籍番号: {列名: 値}}, 警告メッセージのリスト)
    """
    warnings: List[str] = []
    for page in pages:
        if page.unmatched_count:
            warnings.append(
                f"{page.label}（{page.source_path}）: "
                f"学籍番号未確認の行が{page.unmatched_count}件あり、統合対象から除外しました。"
            )

    all_max_scores_known = all(p.max_score is not None for p in pages)
    total_max_score = sum(p.max_score for p in pages) if all_max_scores_known else None

    student_ids: List[str] = []
    seen = set()
    for page in pages:
        for sid in page.rows_by_student_id:
            if sid not in seen:
                seen.add(sid)
                student_ids.append(sid)

    merged: Dict[str, Dict] = {}
    for sid in student_ids:
        row: Dict = {}
        missing_pages = []
        grand_total = 0.0
        any_total = False
        for page in pages:
            page_row = page.rows_by_student_id.get(sid)
            if page_row is None:
                missing_pages.append(page.label)
                for qcol in page.question_columns:
                    row[f"{page.label}: {qcol}"] = None
                row[f"{page.label} 合計"] = None
                continue
            for qcol in page.question_columns:
                row[f"{page.label}: {qcol}"] = page_row.get(qcol)
            page_total = page_row.get(TOTAL_COLUMN)
            row[f"{page.label} 合計"] = page_total
            if isinstance(page_total, (int, float)):
                grand_total += page_total
                any_total = True

        row['総合計'] = grand_total if any_total else None
        row['総配点'] = total_max_score
        row['欠落ページ'] = ", ".join(missing_pages)
        merged[sid] = row

    return merged, warnings


def write_merged_excel(merged_rows: Dict[str, Dict], pages: List[PageSummary], output_path: str) -> None:
    """統合結果をExcelに書き出す。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "複数ページ統合"

    headers = ["No", "学籍番号"]
    for page in pages:
        for qcol in page.question_columns:
            headers.append(f"{page.label}: {qcol}")
        headers.append(f"{page.label} 合計")
    headers += ["総合計", "総配点", "欠落ページ"]

    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    missing_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    for row_idx, (sid, row_data) in enumerate(merged_rows.items(), 2):
        has_missing = bool(row_data.get('欠落ページ'))
        values = [row_idx - 1, sid]
        for page in pages:
            for qcol in page.question_columns:
                values.append(row_data.get(f"{page.label}: {qcol}"))
            values.append(row_data.get(f"{page.label} 合計"))
        values += [row_data.get('総合計'), row_data.get('総配点'), row_data.get('欠落ページ')]

        for col_idx, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.border = thin_border
            cell.alignment = center
            if has_missing:
                cell.fill = missing_fill

    ws.freeze_panes = 'C2'
    ws.column_dimensions['B'].width = 16
    for col_idx in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # 全ページ合計に基づく試験統計
    stats_ws = wb.create_sheet("全ページ試験統計")
    complete_totals = [
        row.get('総合計') for row in merged_rows.values()
        if not row.get('欠落ページ') and isinstance(row.get('総合計'), (int, float))
    ]
    total_max = next(
        (row.get('総配点') for row in merged_rows.values()
         if isinstance(row.get('総配点'), (int, float))),
        None,
    )
    stat_rows = [
        ("受験者数（全ページあり）", len(complete_totals)),
        ("満点", total_max),
        ("平均点", round(statistics.mean(complete_totals), 2) if complete_totals else 0),
        ("標準偏差", round(statistics.stdev(complete_totals), 2) if len(complete_totals) > 1 else 0),
        ("最高点", max(complete_totals) if complete_totals else 0),
        ("最低点", min(complete_totals) if complete_totals else 0),
        ("ページ不足人数", sum(bool(row.get('欠落ページ')) for row in merged_rows.values())),
    ]
    stats_ws.append(["項目", "値"])
    for cell in stats_ws[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
    for item in stat_rows:
        stats_ws.append(item)
    stats_ws.column_dimensions['A'].width = 28
    stats_ws.column_dimensions['B'].width = 16

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def run_multi_page_import_gui(
    image_folder: str,
    parent: Optional[tk.Tk] = None,
    on_project_change=None,
    total_pages: Optional[int] = None,
) -> None:
    """複数ページ答案の試験ページ・取込バッチ管理画面を開く。"""
    from constants import BOXED_FOLDER, RESULTS_DATA_FOLDER, RESULTS_FOLDER

    data_dir = Path(image_folder) / RESULTS_FOLDER / RESULTS_DATA_FOLDER
    managed_root = data_dir / 'multi_page_imports'
    manifest_path = data_dir / MULTI_PAGE_MANIFEST_FILE
    managed_root.mkdir(parents=True, exist_ok=True)

    if manifest_path.exists():
        try:
            batches, saved_audit = load_multi_page_manifest(str(manifest_path))
            expected_pages = list(saved_audit.exam_pages)
            saved_shared_layout = saved_audit.shared_layout
            saved_active_page = saved_audit.active_exam_page
        except Exception as exc:
            messagebox.showerror(
                "複数ページ答案",
                f"保存済みの取込情報を読み込めませんでした:\n{exc}",
                parent=parent,
            )
            return
    else:
        batches = []
        # 新規案件では、ページ番号を都度入力させず、最初に総ページ数だけを
        # 確認して1ページ目から順番に取り込む。
        if total_pages is None:
            total_pages = simpledialog.askinteger(
                "総ページ数", "この答案は全部で何ページですか？",
                parent=parent, minvalue=1,
            )
        if total_pages is None:
            return
        expected_pages = list(range(1, total_pages + 1))
        saved_shared_layout = False
        saved_active_page = None

    window = tk.Toplevel(parent)
    window.title("答案ファイルの追加")
    window.geometry("900x620")
    if parent is not None:
        window.transient(parent)

    header = tk.Frame(window, bg="#37474F")
    header.pack(fill=tk.X)
    tk.Label(
        header,
        text="答案ファイルの追加",
        bg="#37474F", fg="white",
        font=(UI_FONT, get_ui_font_size(12), 'bold'),
    ).pack(anchor=tk.W, padx=12, pady=(10, 2))
    tk.Label(
        header,
        text=(
            "ページを選択して、答案ファイルを追加してください。\n"
            "ファイル選択画面では複数ファイルをまとめて選べます"
            "（⌘/Ctrl+クリックや範囲選択で複数選択）。"
        ),
        bg="#37474F", fg="#ECEFF1", font=(UI_FONT, get_ui_font_size(9)), justify=tk.LEFT,
    ).pack(anchor=tk.W, padx=12, pady=(0, 10))

    shared_layout_var = tk.BooleanVar(value=saved_shared_layout)
    shared_layout_frame = tk.Frame(window, bg="#FFF3E0", padx=12, pady=8)
    shared_layout_frame.pack(fill=tk.X, padx=12, pady=(10, 0))
    tk.Checkbutton(
        shared_layout_frame,
        text="💡 学籍番号欄・氏名欄・解答欄の初期設定を全ページで共通にする",
        variable=shared_layout_var,
        command=lambda: (_save(), _refresh()),
        anchor=tk.W, bg="#FFF3E0", activebackground="#FFF3E0",
        highlightthickness=0, cursor="hand2",
        font=(UI_FONT, get_ui_font_size(10), "bold"), fg="#E65100",
    ).pack(fill=tk.X)
    tk.Label(
        shared_layout_frame,
        text=(
            "オンにすると、1ページ目で設定した学籍番号欄・氏名欄・解答欄の位置を、"
            "2ページ目以降にも自動で適用します。"
        ),
        bg="#FFF3E0", fg="#795548", anchor=tk.W, justify=tk.LEFT,
        font=(UI_FONT, get_ui_font_size(8)), wraplength=650,
    ).pack(fill=tk.X, pady=(2, 0))

    # 左に取込ボタン用サイドバー、右にツリーを横並びで配置する。
    # 固定ピクセルのplace()は、上の案内文の行数（＝高さ）が変わるたびに
    # 位置がずれるため使わない。
    content_frame = tk.Frame(window)
    content_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)

    buttons = tk.Frame(content_frame, bg="#ECEFF1", padx=10, pady=10, width=165)
    buttons.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 12))
    buttons.pack_propagate(False)

    tree_frame = tk.Frame(content_frame)
    tree_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    tree = ttk.Treeview(
        tree_frame,
        columns=('sources', 'answers', 'status'),
        show='tree headings',
        selectmode='browse',
    )
    tree.heading('#0', text='試験ページ（▶ は現在開いているページ）')
    tree.heading('sources', text='取込状況 / PDF・画像')
    tree.heading('answers', text='答案枚数')
    tree.heading('status', text='状態')
    tree.column('#0', width=300, stretch=True)
    tree.column('sources', width=250, stretch=True)
    tree.column('answers', width=80, anchor=tk.CENTER, stretch=False)
    tree.column('status', width=100, anchor=tk.CENTER, stretch=False)
    scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    status_var = tk.StringVar()
    active_page = [saved_active_page]
    status_label = tk.Label(
        window, textvariable=status_var, anchor=tk.W, justify=tk.LEFT,
        font=(UI_FONT, get_ui_font_size(9)), fg="#455A64",
    )
    status_label.pack(fill=tk.X, padx=12, pady=(0, 6))

    def _all_answers():
        return [answer for batch in batches for answer in batch.answer_pages]

    def _save():
        audit = audit_answer_pages(_all_answers(), expected_pages=sorted(expected_pages))
        audit.shared_layout = shared_layout_var.get()
        audit.active_exam_page = active_page[0]
        save_multi_page_manifest(batches, audit, str(manifest_path))
        return audit

    def _refresh(select_iid=None):
        tree.delete(*tree.get_children())
        for page in sorted(expected_pages):
            page_batches = [batch for batch in batches if batch.exam_page == page]
            page_answer_count = sum(len(batch.answer_pages) for batch in page_batches)
            page_iid = f"page:{page}"
            workspace = Path(image_folder) / '_multi_page_pages' / f"page_{page:03d}"
            page_data = workspace / RESULTS_FOLDER / RESULTS_DATA_FOLDER
            if (page_data / 'descriptive_config.json').is_file():
                page_status = '初期設定済み'
            elif (workspace / RESULTS_FOLDER / BOXED_FOLDER).is_dir():
                page_status = '採点準備済み'
            elif page_answer_count:
                page_status = '取込済み'
            else:
                page_status = '未取込'
            current_mark = '▶ ' if active_page[0] == page else ''
            tree.insert(
                '', tk.END, iid=page_iid,
                text=f"{current_mark}試験ページ {page}",
                values=(f"{len(page_batches)}バッチ", page_answer_count, page_status),
                open=True,
            )
            for batch in page_batches:
                names = ', '.join(Path(path).name for path in batch.source_paths)
                batch_iid = f"batch:{batch.batch_id}"
                tree.insert(
                    page_iid, tk.END, iid=batch_iid,
                    text=f"取込 {batch.batch_id[:8]}",
                    values=(names, len(batch.answer_pages), '取込済み'),
                )
                for answer in batch.answer_pages:
                    excluded = getattr(answer, 'status', 'imported') == 'excluded'
                    tree.insert(
                        batch_iid, tk.END, iid=f"answer:{answer.image_id}",
                        text=Path(answer.source_path).name,
                        values=(
                            answer.student_id or '(未確認)',
                            answer.exam_page,
                            '除外中' if excluded else '取込済み',
                        ),
                    )
        audit = audit_answer_pages(_all_answers(), expected_pages=sorted(expected_pages))
        status_var.set(
            f"試験ページ: {len(expected_pages)}　取込バッチ: {len(batches)}　"
            f"答案画像: {len(_all_answers())}　学籍番号未確認: {len(audit.unmatched_image_ids)}"
        )
        if select_iid and tree.exists(select_iid):
            tree.selection_set(select_iid)
            tree.see(select_iid)
        elif not tree.selection() and expected_pages:
            # 何も選択されていなければ、現在の作業ページか先頭のページを
            # 自動選択し、追加先を分かりやすくする。
            default_page = active_page[0] if active_page[0] in expected_pages else sorted(expected_pages)[0]
            tree.selection_set(f"page:{default_page}")

    def _selected_page():
        selection = tree.selection()
        if not selection:
            return None
        iid = selection[0]
        if iid.startswith('page:'):
            return int(iid.split(':', 1)[1])
        if iid.startswith('batch:'):
            batch_id = iid.split(':', 1)[1]
            return next(batch.exam_page for batch in batches if batch.batch_id == batch_id)
        return None

    def _add_page():
        page = simpledialog.askinteger(
            "試験ページを追加", "試験ページ番号:", parent=window, minvalue=1,
        )
        if page is None:
            return
        if page in expected_pages:
            messagebox.showinfo("確認", f"試験ページ {page} は既にあります。", parent=window)
            return
        expected_pages.append(page)
        _save()
        _refresh(f"page:{page}")

    def _add_files():
        page = _selected_page()
        if page is None:
            if not expected_pages:
                messagebox.showwarning("確認", "先に試験ページを追加してください。", parent=window)
                return
            page = simpledialog.askinteger(
                "取込先", "追加する試験ページ番号:", parent=window, minvalue=1,
            )
            if page is None:
                return
        if page not in expected_pages:
            messagebox.showerror("エラー", f"試験ページ {page} は登録されていません。", parent=window)
            return
        paths = list(filedialog.askopenfilenames(
            title=f"試験ページ {page} のPDF・画像を追加（複数選択可）",
            filetypes=[
                ("PDF・画像", "*.pdf *.png *.jpg *.jpeg *.bmp *.tif *.tiff"),
                ("すべてのファイル", "*.*"),
            ],
            parent=window, initialdir=image_folder,
        ))
        if not paths:
            return
        existing = {str(Path(path).resolve()) for batch in batches for path in batch.source_paths}
        repeated = [path for path in paths if str(Path(path).resolve()) in existing]
        if repeated:
            messagebox.showerror(
                "重複する取込元",
                "既に取り込まれているファイルがあります:\n"
                + "\n".join(Path(path).name for path in repeated),
                parent=window,
            )
            return
        try:
            status_var.set("PDF・画像を取り込んでいます…")
            window.update_idletasks()
            batch = import_files_as_batch(page, paths, str(managed_root))
            batches.append(batch)
            _save()
            _refresh(f"batch:{batch.batch_id}")
        except Exception as exc:
            messagebox.showerror("取込エラー", str(exc), parent=window)
            _refresh()

    def _guided_import():
        """ページ順に、ファイル選択ダイアログだけでまとめて取り込む。

        1回の選択で複数ファイルをまとめて選べるため、案内・確認の
        ポップアップは挟まずページごとに1回だけファイル選択を行う。
        後から追加のファイルが必要になった場合は「答案画像ファイルを
        追加」からいつでも追加できる。
        """
        for page in sorted(expected_pages):
            already = any(batch.exam_page == page for batch in batches)
            if already:
                continue
            paths = list(filedialog.askopenfilenames(
                title=f"試験ページ {page} のPDF・画像を選択（複数選択可）",
                filetypes=[
                    ("PDF・画像", "*.pdf *.png *.jpg *.jpeg *.bmp *.tif *.tiff"),
                    ("すべてのファイル", "*.*"),
                ], parent=window, initialdir=image_folder,
            ))
            if not paths:
                continue
            existing = {str(Path(path).resolve()) for batch in batches for path in batch.source_paths}
            repeated = [path for path in paths if str(Path(path).resolve()) in existing]
            if repeated:
                messagebox.showerror("重複する取込元", "既に取り込まれているファイルがあります。", parent=window)
                continue
            try:
                status_var.set(f"試験ページ {page} を取り込んでいます…")
                window.update_idletasks()
                batch = import_files_as_batch(page, paths, str(managed_root))
                batches.append(batch)
                _save()
                _refresh(f"batch:{batch.batch_id}")
            except Exception as exc:
                messagebox.showerror("取込エラー", str(exc), parent=window)
        _refresh()

    def _change_page():
        selection = tree.selection()
        if not selection or not selection[0].startswith('batch:'):
            messagebox.showwarning("確認", "変更する取込バッチを選択してください。", parent=window)
            return
        batch_id = selection[0].split(':', 1)[1]
        batch = next(batch for batch in batches if batch.batch_id == batch_id)
        page = simpledialog.askinteger(
            "試験ページを変更", "変更後の試験ページ番号:",
            initialvalue=batch.exam_page, minvalue=1, parent=window,
        )
        if page is None or page == batch.exam_page:
            return
        if page not in expected_pages:
            expected_pages.append(page)
        batch.exam_page = page
        for answer in batch.answer_pages:
            answer.exam_page = page
        _save()
        _refresh(f"batch:{batch.batch_id}")

    def _delete_selected():
        selection = tree.selection()
        if not selection:
            return
        iid = selection[0]
        if iid.startswith('batch:'):
            batch_id = iid.split(':', 1)[1]
            batch = next(batch for batch in batches if batch.batch_id == batch_id)
            if not messagebox.askyesno(
                "取込バッチを削除",
                f"取込バッチ {batch.batch_id[:8]} と展開画像 {len(batch.answer_pages)}枚を削除しますか？\n"
                "元のPDF・画像は削除されません。",
                parent=window,
            ):
                return
            batch_dir = managed_root / batch.batch_id
            if batch_dir.parent == managed_root and batch_dir.exists():
                shutil.rmtree(batch_dir)
            batches.remove(batch)
            _save()
            _refresh(f"page:{batch.exam_page}")
        elif iid.startswith('page:'):
            page = int(iid.split(':', 1)[1])
            if any(batch.exam_page == page for batch in batches):
                messagebox.showwarning(
                    "削除できません", "先にこの試験ページの取込バッチを削除してください。",
                    parent=window,
                )
                return
            expected_pages.remove(page)
            _save()
            _refresh()

    def _selected_answer():
        selection = tree.selection()
        if not selection or not selection[0].startswith('answer:'):
            return None
        image_id = selection[0].split(':', 1)[1]
        return next(
            (a for b in batches for a in b.answer_pages if a.image_id == image_id), None,
        )

    def _change_answer_exam_page():
        answer = _selected_answer()
        if answer is None:
            messagebox.showwarning("確認", "変更する答案を選択してください。", parent=window)
            return
        page = simpledialog.askinteger(
            "この答案だけページを変更", "変更後の試験ページ番号:",
            initialvalue=answer.exam_page, minvalue=1, parent=window,
        )
        if page is None or page == answer.exam_page:
            return
        if page not in expected_pages:
            expected_pages.append(page)
        answer.exam_page = page
        _save()
        _refresh(f"answer:{answer.image_id}")

    def _toggle_answer_excluded():
        answer = _selected_answer()
        if answer is None:
            messagebox.showwarning("確認", "対象の答案を選択してください。", parent=window)
            return
        if getattr(answer, 'status', 'imported') != 'excluded':
            if not messagebox.askyesno(
                "答案を除外",
                "採点・集計の対象から除外しますか？\n画像ファイルは削除されません。",
                parent=window,
            ):
                return
        toggle_answer_status(batches, answer.image_id)
        _save()
        _refresh(f"answer:{answer.image_id}")

    def _change_page_dispatch():
        selection = tree.selection()
        if selection and selection[0].startswith('answer:'):
            _change_answer_exam_page()
        else:
            _change_page()

    def _delete_dispatch():
        selection = tree.selection()
        if selection and selection[0].startswith('answer:'):
            _toggle_answer_excluded()
        else:
            _delete_selected()

    if not manifest_path.exists():
        window.after(50, _guided_import)
    tk.Button(buttons, text="答案画像ファイルを追加", command=_add_files,
              bg="#90CAF9", fg="#263238", font=(UI_FONT, get_ui_font_size(11), 'bold'),
              height=3, wraplength=135).pack(fill=tk.X, pady=(0, 10))
    tk.Button(
        buttons, text="ページを変更", command=_change_page_dispatch,
        bg="#ECEFF1", fg="#37474F", font=(UI_FONT, get_ui_font_size(9), 'bold'),
        wraplength=135,
    ).pack(fill=tk.X, pady=(0, 4))
    tk.Button(
        buttons, text="削除／除外", command=_delete_dispatch,
        bg="#ECEFF1", fg="#37474F", font=(UI_FONT, get_ui_font_size(9), 'bold'),
        wraplength=135,
    ).pack(fill=tk.X, pady=(0, 10))
    tk.Button(
        buttons, text="監査結果を確認",
        command=lambda: show_multi_page_audit_dialog(window, image_folder, on_change=_refresh),
        bg="#FFE082", fg="#5D4037", font=(UI_FONT, get_ui_font_size(9), 'bold'),
        wraplength=135,
    ).pack(fill=tk.X, pady=(0, 10))

    def _close():
        try:
            _save()
        except Exception as exc:
            messagebox.showerror("保存エラー", str(exc), parent=window)
            return
        window.destroy()
        if on_project_change is not None:
            on_project_change()

    tk.Button(
        window, text="保存して閉じる", command=_close,
        bg="#90CAF9", fg="#263238",
        activebackground="#64B5F6", activeforeground="#263238",
        font=(UI_FONT, get_ui_font_size(10), 'bold'),
    ).pack(fill=tk.X, padx=12, pady=(0, 12))
    window.protocol("WM_DELETE_WINDOW", _close)
    _refresh()


def show_multi_page_audit_dialog(parent, project_folder: str, on_change=None) -> None:
    """複数ページ答案の監査結果（不足・重複・未確認・想定外）を一覧表示する。

    問題のある答案を選択して「除外／除外解除」で採点・集計の対象から外せる。
    """
    from constants import RESULTS_DATA_FOLDER, RESULTS_FOLDER

    manifest_path = Path(project_folder) / RESULTS_FOLDER / RESULTS_DATA_FOLDER / MULTI_PAGE_MANIFEST_FILE
    if not manifest_path.is_file():
        messagebox.showinfo("監査結果", "複数ページ答案が設定されていません。", parent=parent)
        return
    batches, saved_audit = load_multi_page_manifest(str(manifest_path))

    window = tk.Toplevel(parent)
    window.title("監査結果を確認")

    tk.Label(
        window,
        text=(
            "ページ不足・重複・学籍番号未確認・想定外の答案を一覧表示します。\n"
            "答案を選択して「除外／除外解除」を押すと、採点・集計の対象から外せます。"
        ),
        justify=tk.LEFT, anchor=tk.W, font=(UI_FONT, get_ui_font_size(9)),
    ).pack(fill=tk.X, padx=12, pady=(10, 6))

    tree = ttk.Treeview(window, columns=('detail',), show='tree headings')
    tree.heading('#0', text='カテゴリ / 答案')
    tree.heading('detail', text='詳細')
    tree.column('#0', width=280)
    tree.column('detail', width=380)
    tree.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 6))

    status_var = tk.StringVar()
    tk.Label(
        window, textvariable=status_var, anchor=tk.W,
        font=(UI_FONT, get_ui_font_size(9)), fg="#455A64",
    ).pack(fill=tk.X, padx=12)

    def _answer_by_id(image_id):
        return next(
            (a for b in batches for a in b.answer_pages if a.image_id == image_id), None,
        )

    def _current_audit():
        return audit_answer_pages(
            [a for b in batches for a in b.answer_pages], expected_pages=saved_audit.exam_pages,
        )

    def _refresh():
        tree.delete(*tree.get_children())
        audit = _current_audit()

        missing_node = tree.insert('', tk.END, text="ページ不足", open=True)
        for student_id, pages in audit.missing_pages.items():
            tree.insert(
                missing_node, tk.END,
                text=student_id, values=(f"不足ページ: {', '.join(map(str, pages))}",),
            )

        dup_node = tree.insert('', tk.END, text="同一ページ内の重複", open=True)
        for page, by_student in audit.duplicates.items():
            for student_id, image_ids in by_student.items():
                for image_id in image_ids:
                    answer = _answer_by_id(image_id)
                    detail = f"ページ{page} / 学籍番号{student_id}"
                    if answer:
                        detail += f" / {Path(answer.source_path).name}"
                    tree.insert(
                        dup_node, tk.END, iid=f"answer:{image_id}",
                        text=image_id[:8], values=(detail,),
                    )

        unmatched_node = tree.insert('', tk.END, text="学籍番号未確認", open=True)
        for image_id in audit.unmatched_image_ids:
            answer = _answer_by_id(image_id)
            detail = (
                f"試験ページ{answer.exam_page} / {Path(answer.source_path).name}"
                if answer else ""
            )
            tree.insert(
                unmatched_node, tk.END, iid=f"answer:{image_id}",
                text=image_id[:8], values=(detail,),
            )

        unexpected_node = tree.insert('', tk.END, text="名簿外ID・想定外ページ", open=True)
        for student_id in audit.unexpected_student_ids:
            tree.insert(unexpected_node, tk.END, text=f"名簿外ID: {student_id}", values=("",))
        for image_id, page in audit.unexpected_exam_pages.items():
            answer = _answer_by_id(image_id)
            detail = (
                f"想定外ページ{page} / {Path(answer.source_path).name}"
                if answer else f"想定外ページ{page}"
            )
            tree.insert(
                unexpected_node, tk.END, iid=f"answer:{image_id}",
                text=image_id[:8], values=(detail,),
            )

        status_var.set(
            f"ページ不足: {len(audit.missing_pages)}件　"
            f"重複: {sum(len(v) for v in audit.duplicates.values())}件　"
            f"学籍番号未確認: {len(audit.unmatched_image_ids)}件　"
            f"名簿外/想定外: {len(audit.unexpected_student_ids) + len(audit.unexpected_exam_pages)}件"
        )

    def _toggle_selected():
        selection = tree.selection()
        if not selection or not selection[0].startswith('answer:'):
            messagebox.showwarning("確認", "対象の答案を選択してください。", parent=window)
            return
        image_id = selection[0].split(':', 1)[1]
        answer = _answer_by_id(image_id)
        if answer is None:
            return
        if getattr(answer, 'status', 'imported') != 'excluded':
            if not messagebox.askyesno(
                "答案を除外",
                "採点・集計の対象から除外しますか？\n画像ファイルは削除されません。",
                parent=window,
            ):
                return
        toggle_answer_status(batches, image_id)
        new_audit = _current_audit()
        new_audit.shared_layout = saved_audit.shared_layout
        new_audit.active_exam_page = saved_audit.active_exam_page
        save_multi_page_manifest(batches, new_audit, str(manifest_path))
        _refresh()
        if on_change is not None:
            on_change()

    btn_frame = tk.Frame(window)
    btn_frame.pack(fill=tk.X, padx=12, pady=(0, 12))
    tk.Button(
        btn_frame, text="選択した答案を除外／除外解除", command=_toggle_selected,
        bg="#FFCCBC", fg="#3E2723", font=(UI_FONT, get_ui_font_size(9), 'bold'),
    ).pack(side=tk.LEFT)
    tk.Button(btn_frame, text="閉じる", command=window.destroy).pack(side=tk.RIGHT)

    _refresh()
    fit_window_to_content(window, min_width=700, min_height=500)


def run_multi_page_merge_gui(parent: Optional[tk.Tk] = None) -> Optional[str]:
    """複数ページ統合ツールのGUI本体。

    Returns:
        統合Excelの保存先パス。キャンセル時は None。
    """
    window = tk.Toplevel(parent)
    window.title("複数ページ統合")

    tk.Label(
        window,
        text="ページ順に集計Excelを追加してください\n（例: ページ1→ページ2→ページ3 の順）",
        font=(UI_FONT, get_ui_font_size(10), 'bold'), justify=tk.LEFT,
    ).pack(pady=(10, 5))

    list_frame = tk.Frame(window)
    list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

    listbox = tk.Listbox(list_frame, font=(UI_FONT, get_ui_font_size(9)))
    scrollbar = tk.Scrollbar(list_frame, orient=tk.VERTICAL, command=listbox.yview)
    listbox.configure(yscrollcommand=scrollbar.set)
    listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    file_paths: List[str] = []

    def _refresh_listbox():
        listbox.delete(0, tk.END)
        for i, path in enumerate(file_paths, 1):
            listbox.insert(tk.END, f"ページ{i}: {Path(path).name}")

    def _add_file():
        path = filedialog.askopenfilename(
            title="集計Excelを選択",
            filetypes=[("Excelファイル", "*.xlsx *.xls"), ("すべてのファイル", "*.*")],
            parent=window,
        )
        if path:
            file_paths.append(path)
            _refresh_listbox()

    def _remove_selected():
        sel = listbox.curselection()
        if not sel:
            return
        del file_paths[sel[0]]
        _refresh_listbox()

    def _move(offset):
        sel = listbox.curselection()
        if not sel:
            return
        i = sel[0]
        j = i + offset
        if 0 <= j < len(file_paths):
            file_paths[i], file_paths[j] = file_paths[j], file_paths[i]
            _refresh_listbox()
            listbox.selection_set(j)

    btn_row = tk.Frame(window)
    btn_row.pack(pady=5)
    tk.Button(btn_row, text="＋ 追加", command=_add_file).pack(side=tk.LEFT, padx=3)
    tk.Button(btn_row, text="↑", command=lambda: _move(-1), width=3).pack(side=tk.LEFT, padx=3)
    tk.Button(btn_row, text="↓", command=lambda: _move(1), width=3).pack(side=tk.LEFT, padx=3)
    tk.Button(btn_row, text="－ 削除", command=_remove_selected).pack(side=tk.LEFT, padx=3)

    result_path = [None]

    def _run_merge():
        if len(file_paths) < 2:
            messagebox.showwarning("エラー", "2件以上のExcelを追加してください。", parent=window)
            return

        pages = []
        for i, path in enumerate(file_paths, 1):
            try:
                pages.append(read_page_summary(path, f"ページ{i}"))
            except ValueError as e:
                messagebox.showerror("エラー", str(e), parent=window)
                return
            except Exception as e:
                messagebox.showerror("エラー", f"読み込みに失敗しました:\n{path}\n{e}", parent=window)
                return

        merged, warnings = merge_page_summaries(pages)

        output_path = filedialog.asksaveasfilename(
            title="統合結果の保存先",
            defaultextension=".xlsx",
            filetypes=[("Excelファイル", "*.xlsx")],
            initialfile="複数ページ統合結果.xlsx",
            parent=window,
        )
        if not output_path:
            return

        try:
            write_merged_excel(merged, pages, output_path)
        except Exception as e:
            messagebox.showerror("エラー", f"書き出しに失敗しました:\n{e}", parent=window)
            return

        result_path[0] = output_path
        msg = f"{len(merged)}人分を統合しました。\n保存先: {output_path}"
        if warnings:
            msg += "\n\n【警告】\n" + "\n".join(warnings)
        messagebox.showinfo("完了", msg, parent=window)
        window.destroy()

    tk.Button(
        window, text="統合実行", command=_run_merge,
        bg="#4CAF50", fg="black", font=(UI_FONT, get_ui_font_size(11), 'bold'), height=2,
    ).pack(fill=tk.X, padx=10, pady=10)

    fit_window_to_content(window, min_width=520, min_height=420)
    window.grab_set()
    window.wait_window()

    return result_path[0]
